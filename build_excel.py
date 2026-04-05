# build_excel.py — REIT Rental Analysis Workbook Builder
#
# SETUP (one-time, on any machine):
#   pip install requests pandas openpyxl
#
# USAGE:
#   1. Set GITHUB_TOKEN below to your fine-grained PAT (read scope: Public-CRE-Data/residata)
#   2. Run: python build_excel.py
#   3. Find output: REIT_Rental_Analysis_{date}.xlsx in same folder as this script
#
# WEEKLY WORKFLOW:
#   - Scraper runs on main machine → pushes new dated CSVs to GitHub
#   - Copy this script to any machine with Python → run to get latest Excel
#   - Same-property WoW analysis activates automatically from Week 2 onwards

import os
import io
import json
import re
import sys
from datetime import datetime, timedelta
import requests
import pandas as pd
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.workbook.properties import CalcProperties

# Ensure stdout handles Unicode on Windows (cp1252 terminals)
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf-8-sig"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — edit these before running
# ─────────────────────────────────────────────────────────────────────────────
GITHUB_OWNER  = "Public-CRE-Data"
GITHUB_REPO   = "residata"
# Set GITHUB_TOKEN to your fine-grained PAT (read-only access to this repo)
GITHUB_TOKEN  = ""
DATA_PATH     = "data/raw"          # folder in repo where CSVs live
REGISTRY_PATH = "data/registry/unit_registry.csv"
SUMMARY_PATH  = "data/summary/summary_history.csv"  # persistent summary in repo
OUTPUT_DIR    = "."  # saves Excel in same folder as this script; change to any local path
CACHE_DIR     = "./residata_cache"  # local cache to avoid re-downloading unchanged files
LOCAL_SUMMARY_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "summary")

# ─────────────────────────────────────────────────────────────────────────────
# MACRO MARKET MAP
# ─────────────────────────────────────────────────────────────────────────────
MACRO_MAP = {
    # Seattle
    "Seattle": "Seattle", "Bellevue": "Seattle", "Redmond": "Seattle", "Kirkland": "Seattle",
    "Bothell": "Seattle", "Everett": "Seattle", "Renton": "Seattle", "Kent": "Seattle",
    "Federal Way": "Seattle", "Shoreline": "Seattle", "Lynnwood": "Seattle",
    "Marysville": "Seattle", "Issaquah": "Seattle", "Sammamish": "Seattle",
    "Burien": "Seattle", "Covington WA": "Seattle", "Des Moines WA": "Seattle",
    "Edmonds WA": "Seattle", "Monroe WA": "Seattle", "Mountlake Terrace": "Seattle",
    "Mukilteo": "Seattle", "Puyallup": "Seattle", "SeaTac": "Seattle",
    "Snohomish": "Seattle", "Tacoma": "Seattle",

    # Atlanta
    "Atlanta": "Atlanta", "Alpharetta": "Atlanta", "Marietta": "Atlanta",
    "Sandy Springs": "Atlanta", "Smyrna GA": "Atlanta", "Smyrna": "Atlanta",
    "Kennesaw": "Atlanta", "Roswell GA": "Atlanta", "Roswell": "Atlanta",
    "Decatur": "Atlanta", "Duluth GA": "Atlanta", "Duluth": "Atlanta",
    "Norcross": "Atlanta", "Peachtree City": "Atlanta", "Lawrenceville": "Atlanta",
    "Acworth GA": "Atlanta", "Acworth": "Atlanta", "Auburn GA": "Atlanta",
    "Buford": "Atlanta", "Canton GA": "Atlanta", "Carrollton GA": "Atlanta",
    "Chamblee": "Atlanta", "College Park GA": "Atlanta", "Conyers": "Atlanta",
    "Covington GA": "Atlanta", "Cumming": "Atlanta", "Dacula": "Atlanta",
    "Dallas GA": "Atlanta", "Douglasville": "Atlanta", "Dunwoody": "Atlanta",
    "East Point": "Atlanta", "Fayetteville GA": "Atlanta", "Forest Park": "Atlanta",
    "Griffin GA": "Atlanta", "Hiram": "Atlanta", "Johns Creek": "Atlanta",
    "Jonesboro GA": "Atlanta", "Lilburn": "Atlanta", "Lithia Springs": "Atlanta",
    "Loganville": "Atlanta", "Mableton": "Atlanta", "McDonough": "Atlanta",
    "Newnan": "Atlanta", "Powder Springs": "Atlanta", "Riverdale GA": "Atlanta",
    "Rome GA": "Atlanta", "Snellville": "Atlanta", "Stockbridge GA": "Atlanta",
    "Stone Mountain": "Atlanta", "Sugar Hill GA": "Atlanta", "Suwanee": "Atlanta",
    "Tucker GA": "Atlanta", "Union City GA": "Atlanta", "Winder GA": "Atlanta",
    "Woodstock GA": "Atlanta",

    # Miami/Fort Lauderdale
    "Miami": "Miami/Fort Lauderdale", "Fort Lauderdale": "Miami/Fort Lauderdale",
    "Boca Raton": "Miami/Fort Lauderdale", "Dania Beach": "Miami/Fort Lauderdale",
    "Davie": "Miami/Fort Lauderdale", "Hollywood FL": "Miami/Fort Lauderdale",
    "Hollywood": "Miami/Fort Lauderdale", "Plantation FL": "Miami/Fort Lauderdale",
    "Plantation": "Miami/Fort Lauderdale", "Pembroke Pines": "Miami/Fort Lauderdale",
    "Sunrise FL": "Miami/Fort Lauderdale", "Sunrise": "Miami/Fort Lauderdale",
    "Coral Springs": "Miami/Fort Lauderdale", "Aventura": "Miami/Fort Lauderdale",
    "Hallandale Beach": "Miami/Fort Lauderdale", "Hialeah": "Miami/Fort Lauderdale",
    "Homestead": "Miami/Fort Lauderdale", "Kendall FL": "Miami/Fort Lauderdale",
    "Margate FL": "Miami/Fort Lauderdale", "Margate": "Miami/Fort Lauderdale",
    "Miami Beach": "Miami/Fort Lauderdale", "Miami Gardens": "Miami/Fort Lauderdale",
    "Miramar FL": "Miami/Fort Lauderdale", "Miramar": "Miami/Fort Lauderdale",
    "North Lauderdale": "Miami/Fort Lauderdale", "North Miami": "Miami/Fort Lauderdale",
    "Oakland Park FL": "Miami/Fort Lauderdale", "Oakland Park": "Miami/Fort Lauderdale",
    "Boynton Beach": "Miami/Fort Lauderdale", "Delray Beach": "Miami/Fort Lauderdale",
    "Pompano Beach": "Miami/Fort Lauderdale", "Tamarac": "Miami/Fort Lauderdale",
    "West Palm Beach": "Miami/Fort Lauderdale", "Weston FL": "Miami/Fort Lauderdale",
    "Weston": "Miami/Fort Lauderdale",

    # Orlando
    "Orlando": "Orlando", "Winter Park FL": "Orlando", "Winter Park": "Orlando",
    "Kissimmee": "Orlando", "Ocoee": "Orlando", "Altamonte Springs": "Orlando",
    "Lake Mary FL": "Orlando", "Lake Mary": "Orlando", "Sanford FL": "Orlando",
    "Sanford": "Orlando", "Casselberry": "Orlando", "Apopka": "Orlando",
    "Clermont FL": "Orlando", "Daytona Beach": "Orlando", "Deltona": "Orlando",
    "Longwood FL": "Orlando", "Longwood": "Orlando", "Melbourne FL": "Orlando",
    "Orange City FL": "Orlando", "Oviedo": "Orlando", "Palm Bay": "Orlando",
    "Poinciana": "Orlando", "Port Orange": "Orlando", "St Cloud FL": "Orlando",
    "Winter Garden": "Orlando", "Winter Haven": "Orlando", "Winter Springs": "Orlando",
    "Osceola County": "Orlando",

    # Tampa
    "Tampa": "Tampa", "St. Petersburg": "Tampa", "Clearwater": "Tampa",
    "Wesley Chapel": "Tampa", "Sarasota": "Tampa", "Brandon FL": "Tampa",
    "Brandon": "Tampa", "Riverview FL": "Tampa", "Riverview": "Tampa",
    "Land O Lakes": "Tampa", "Bradenton": "Tampa", "Brooksville FL": "Tampa",
    "Dunedin FL": "Tampa", "Hernando County FL": "Tampa", "Largo FL": "Tampa",
    "Largo": "Tampa", "Lutz FL": "Tampa", "Lutz": "Tampa",
    "New Port Richey": "Tampa", "Nokomis": "Tampa", "Palmetto FL": "Tampa",
    "Plant City": "Tampa", "Ruskin FL": "Tampa", "Seffner": "Tampa",
    "Spring Hill FL": "Tampa", "Spring Hill": "Tampa", "St. Pete Beach": "Tampa",
    "Sun City Center": "Tampa", "Temple Terrace": "Tampa", "Venice FL": "Tampa",
    "Wimauma": "Tampa", "Zephyrhills": "Tampa", "Lakeland": "Tampa",

    # Denver
    "Denver": "Denver", "Aurora CO": "Denver", "Aurora": "Denver",
    "Westminster CO": "Denver", "Westminster": "Denver", "Englewood CO": "Denver",
    "Englewood": "Denver", "Lakewood CO": "Denver", "Lakewood": "Denver",
    "Thornton": "Denver", "Arvada": "Denver", "Broomfield": "Denver",
    "Brighton CO": "Denver", "Littleton CO": "Denver", "Littleton": "Denver",
    "Centennial CO": "Denver", "Centennial": "Denver", "Parker CO": "Denver",
    "Parker": "Denver", "Northglenn": "Denver", "Highlands Ranch": "Denver",
    "Castle Rock CO": "Denver", "Commerce City CO": "Denver", "Erie CO": "Denver",
    "Firestone": "Denver", "Fort Collins": "Denver", "Boulder CO": "Denver",
    "Longmont": "Denver", "Louisville CO": "Denver", "Wheat Ridge": "Denver",

    # Raleigh/Durham
    "Raleigh": "Raleigh/Durham", "Durham": "Raleigh/Durham", "Cary": "Raleigh/Durham",
    "Chapel Hill": "Raleigh/Durham", "Morrisville": "Raleigh/Durham",
    "Apex": "Raleigh/Durham", "Wake Forest": "Raleigh/Durham",
    "Garner": "Raleigh/Durham", "Clayton NC": "Raleigh/Durham",
    "Carrboro": "Raleigh/Durham", "Fuquay-Varina": "Raleigh/Durham",
    "Holly Springs NC": "Raleigh/Durham", "Knightdale": "Raleigh/Durham",
    "Zebulon": "Raleigh/Durham",

    # Charlotte
    "Charlotte": "Charlotte", "Concord NC": "Charlotte", "Concord": "Charlotte",
    "Matthews": "Charlotte", "Huntersville": "Charlotte", "Mooresville": "Charlotte",
    "Gastonia": "Charlotte", "Mint Hill": "Charlotte", "Cornelius NC": "Charlotte",
    "Harrisburg NC": "Charlotte", "Indian Trail": "Charlotte",
    "Kannapolis": "Charlotte", "Monroe NC": "Charlotte", "Stallings": "Charlotte",
    "Ballantyne": "Charlotte",

    # San Diego
    "San Diego": "San Diego", "Carlsbad": "San Diego", "Chula Vista": "San Diego",
    "Oceanside": "San Diego", "Escondido": "San Diego", "El Cajon": "San Diego",
    "La Mesa": "San Diego", "Santee": "San Diego", "Encinitas": "San Diego",
    "Alpine CA": "San Diego", "Alpine": "San Diego", "Bonita CA": "San Diego",
    "Lemon Grove": "San Diego", "National City": "San Diego", "Poway": "San Diego",
    "Ramona": "San Diego", "San Marcos CA": "San Diego", "San Marcos": "San Diego",
    "Spring Valley CA": "San Diego", "Vista CA": "San Diego", "Vista": "San Diego",

    # Los Angeles
    "Los Angeles": "Los Angeles", "Long Beach": "Los Angeles",
    "Glendale": "Los Angeles", "Burbank": "Los Angeles", "Pasadena": "Los Angeles",
    "West Hollywood": "Los Angeles", "Culver City": "Los Angeles",
    "El Segundo": "Los Angeles", "Torrance": "Los Angeles", "Hawthorne": "Los Angeles",
    "Inglewood": "Los Angeles", "Compton": "Los Angeles", "Carson": "Los Angeles",
    "Alhambra": "Los Angeles", "Arcadia CA": "Los Angeles", "Arcadia": "Los Angeles",
    "Azusa": "Los Angeles", "Baldwin Park": "Los Angeles", "Bell": "Los Angeles",
    "Bellflower": "Los Angeles", "Canoga Park": "Los Angeles", "Cerritos": "Los Angeles",
    "Chatsworth": "Los Angeles", "Covina": "Los Angeles", "Diamond Bar": "Los Angeles",
    "Downey": "Los Angeles", "Duarte": "Los Angeles", "El Monte": "Los Angeles",
    "Encino": "Los Angeles", "Gardena": "Los Angeles", "Glendora": "Los Angeles",
    "Granada Hills": "Los Angeles", "Hacienda Heights": "Los Angeles",
    "La Puente": "Los Angeles", "Lakewood CA": "Los Angeles", "Lancaster CA": "Los Angeles",
    "Lawndale": "Los Angeles", "Lomita": "Los Angeles", "Monrovia CA": "Los Angeles",
    "Montebello": "Los Angeles", "Monterey Park": "Los Angeles",
    "Northridge": "Los Angeles", "Norwalk CA": "Los Angeles", "Paramount CA": "Los Angeles",
    "Pomona": "Los Angeles", "Reseda": "Los Angeles", "Rowland Heights": "Los Angeles",
    "San Dimas": "Los Angeles", "San Gabriel": "Los Angeles", "San Pedro": "Los Angeles",
    "Santa Clarita": "Los Angeles", "Santa Monica": "Los Angeles",
    "Sherman Oaks": "Los Angeles", "Signal Hill": "Los Angeles",
    "Simi Valley": "Los Angeles", "South Gate": "Los Angeles",
    "Temple City": "Los Angeles", "Thousand Oaks": "Los Angeles",
    "Van Nuys": "Los Angeles", "West Covina": "Los Angeles", "Whittier": "Los Angeles",
    "Woodland Hills": "Los Angeles",

    # Austin
    "Austin": "Austin", "Round Rock": "Austin", "Cedar Park": "Austin",
    "Georgetown TX": "Austin", "Georgetown": "Austin", "Pflugerville": "Austin",
    "Kyle TX": "Austin", "Kyle": "Austin", "Buda": "Austin", "Leander": "Austin",
    "Manor TX": "Austin", "Lakeway": "Austin", "Bastrop": "Austin",
    "Waco": "Austin",

    # Nashville
    "Nashville": "Nashville", "Brentwood TN": "Nashville", "Brentwood": "Nashville",
    "Franklin TN": "Nashville", "Franklin": "Nashville",
    "Murfreesboro": "Nashville", "Smyrna TN": "Nashville",
    "Antioch TN": "Nashville", "Antioch": "Nashville",
    "Clarksville TN": "Nashville", "Columbia TN": "Nashville",
    "Gallatin": "Nashville", "Goodlettsville": "Nashville",
    "Hendersonville TN": "Nashville", "Hendersonville": "Nashville",
    "Hermitage": "Nashville", "Jackson TN": "Nashville",
    "LaVergne": "Nashville", "Mount Juliet": "Nashville",
    "Nolensville": "Nashville", "Old Hickory": "Nashville",
    "Spring Hill TN": "Nashville",

    # Houston
    "Houston": "Houston", "Sugar Land": "Houston", "The Woodlands": "Houston",
    "Katy": "Houston", "Spring TX": "Houston", "Spring": "Houston",
    "Pearland": "Houston", "Pasadena TX": "Houston", "Conroe": "Houston",
    "Humble": "Houston", "Missouri City": "Houston", "Cypress TX": "Houston",
    "Cypress": "Houston", "Deer Park TX": "Houston", "Friendswood": "Houston",
    "Galveston": "Houston", "La Marque": "Houston", "League City": "Houston",
    "Rosenberg TX": "Houston", "Stafford TX": "Houston", "Texas City": "Houston",
    "Webster TX": "Houston",

    # Washington, DC
    "Washington DC": "Washington, DC", "Washington, DC": "Washington, DC",
    "Arlington VA": "Washington, DC", "Arlington": "Washington, DC",
    "Alexandria VA": "Washington, DC", "Alexandria": "Washington, DC",
    "Bethesda MD": "Washington, DC", "Bethesda": "Washington, DC",
    "Silver Spring MD": "Washington, DC", "Silver Spring": "Washington, DC",
    "Rockville MD": "Washington, DC", "Rockville": "Washington, DC",
    "Gaithersburg MD": "Washington, DC", "Gaithersburg": "Washington, DC",
    "Reston VA": "Washington, DC", "Reston": "Washington, DC",
    "McLean VA": "Washington, DC", "McLean": "Washington, DC",
    "Falls Church VA": "Washington, DC", "Falls Church": "Washington, DC",
    "Fairfax VA": "Washington, DC", "Fairfax": "Washington, DC",
    "Herndon VA": "Washington, DC", "Herndon": "Washington, DC",
    "Sterling VA": "Washington, DC", "Sterling": "Washington, DC",
    "Tysons": "Washington, DC", "Centreville VA": "Washington, DC",
    "Centreville": "Washington, DC", "Chantilly": "Washington, DC",
    "Leesburg VA": "Washington, DC", "Leesburg": "Washington, DC",
    "Manassas": "Washington, DC", "Stafford VA": "Washington, DC",
    "Fredericksburg VA": "Washington, DC",

    # Dallas/Fort Worth
    "Dallas": "Dallas/Fort Worth", "Fort Worth": "Dallas/Fort Worth",
    "Plano": "Dallas/Fort Worth", "Irving": "Dallas/Fort Worth",
    "Frisco": "Dallas/Fort Worth", "McKinney": "Dallas/Fort Worth",
    "Garland": "Dallas/Fort Worth", "Mesquite": "Dallas/Fort Worth",
    "Arlington TX": "Dallas/Fort Worth", "Arlington": "Dallas/Fort Worth",
    "Carrollton": "Dallas/Fort Worth", "Richardson": "Dallas/Fort Worth",
    "Grand Prairie": "Dallas/Fort Worth", "Lewisville": "Dallas/Fort Worth",
    "Denton": "Dallas/Fort Worth", "Flower Mound": "Dallas/Fort Worth",
    "Allen TX": "Dallas/Fort Worth", "Allen": "Dallas/Fort Worth",
    "Murphy TX": "Dallas/Fort Worth", "Murphy": "Dallas/Fort Worth",
    "Addison TX": "Dallas/Fort Worth", "Addison": "Dallas/Fort Worth",
    "Balch Springs": "Dallas/Fort Worth", "Cedar Hill": "Dallas/Fort Worth",
    "Coppell": "Dallas/Fort Worth", "DeSoto TX": "Dallas/Fort Worth",
    "Duncanville": "Dallas/Fort Worth", "Euless": "Dallas/Fort Worth",
    "Farmers Branch": "Dallas/Fort Worth", "Grapevine": "Dallas/Fort Worth",
    "Haltom City": "Dallas/Fort Worth", "Lancaster TX": "Dallas/Fort Worth",
    "Little Elm": "Dallas/Fort Worth", "Mansfield TX": "Dallas/Fort Worth",
    "North Richland Hills": "Dallas/Fort Worth", "Prosper": "Dallas/Fort Worth",
    "Rowlett": "Dallas/Fort Worth", "Sachse": "Dallas/Fort Worth",
    "Sunnyvale TX": "Dallas/Fort Worth", "The Colony": "Dallas/Fort Worth",
    "Trophy Club": "Dallas/Fort Worth", "Waxahachie": "Dallas/Fort Worth",
    "Wylie": "Dallas/Fort Worth",

    # Phoenix
    "Phoenix": "Phoenix", "Scottsdale": "Phoenix", "Tempe": "Phoenix",
    "Chandler": "Phoenix", "Mesa": "Phoenix", "Gilbert AZ": "Phoenix",
    "Gilbert": "Phoenix", "Glendale AZ": "Phoenix",
    "Peoria AZ": "Phoenix", "Peoria": "Phoenix",
    "Surprise AZ": "Phoenix", "Surprise": "Phoenix",
    "Goodyear": "Phoenix", "Avondale": "Phoenix", "Buckeye": "Phoenix",
    "Queen Creek": "Phoenix", "Maricopa AZ": "Phoenix",
    "Maricopa": "Phoenix", "Tolleson": "Phoenix",

    # San Francisco-East Bay
    "Oakland": "San Francisco-East Bay", "Berkeley": "San Francisco-East Bay",
    "Fremont CA": "San Francisco-East Bay", "Fremont": "San Francisco-East Bay",
    "Hayward": "San Francisco-East Bay", "Concord CA": "San Francisco-East Bay",
    "Walnut Creek": "San Francisco-East Bay", "Pleasanton": "San Francisco-East Bay",
    "Livermore": "San Francisco-East Bay", "San Leandro": "San Francisco-East Bay",
    "Richmond CA": "San Francisco-East Bay", "Alameda": "San Francisco-East Bay",
    "Castro Valley": "San Francisco-East Bay", "Newark CA": "San Francisco-East Bay",
    "San Ramon": "San Francisco-East Bay", "Union City CA": "San Francisco-East Bay",
    "San Lorenzo": "San Francisco-East Bay",

    # San Francisco
    "San Francisco": "San Francisco", "Daly City": "San Francisco",
    "South San Francisco": "San Francisco", "Pacifica": "San Francisco",
    "San Bruno": "San Francisco",

    # San Jose
    "San Jose": "San Jose", "Santa Clara": "San Jose", "Sunnyvale": "San Jose",
    "Cupertino": "San Jose", "Mountain View": "San Jose", "Milpitas": "San Jose",
    "Campbell CA": "San Jose", "Campbell": "San Jose",
    "Los Altos": "San Jose", "Saratoga CA": "San Jose", "Saratoga": "San Jose",
    "Los Gatos": "San Jose",

    # Boston
    "Boston": "Boston", "Cambridge MA": "Boston", "Cambridge": "Boston",
    "Somerville": "Boston", "Quincy": "Boston", "Brookline": "Boston",
    "Newton": "Boston", "Waltham": "Boston", "Watertown MA": "Boston",
    "Watertown": "Boston", "Malden": "Boston", "Medford": "Boston",
    "Woburn": "Boston", "Burlington MA": "Boston", "Burlington": "Boston",
    "Lexington MA": "Boston", "Lexington": "Boston", "Needham": "Boston",
    "Acton": "Boston", "Andover": "Boston", "Bedford MA": "Boston",
    "Beverly": "Boston", "Canton MA": "Boston", "Chelmsford": "Boston",
    "Danvers": "Boston", "Dracut": "Boston", "Framingham": "Boston",
    "Franklin MA": "Boston", "Haverhill": "Boston", "Holliston": "Boston",
    "Hopkinton": "Boston", "Lowell": "Boston", "Lynn": "Boston",
    "Lynnfield": "Boston", "Marlborough": "Boston", "Methuen": "Boston",
    "Milford MA": "Boston", "Natick": "Boston", "North Andover": "Boston",
    "Norwood": "Boston", "Peabody": "Boston", "Reading MA": "Boston",
    "Reading": "Boston", "Stoughton": "Boston", "Sudbury": "Boston",
    "Tewksbury": "Boston", "Tyngsborough": "Boston", "Wakefield MA": "Boston",
    "Wakefield": "Boston", "Wellesley": "Boston", "Westborough": "Boston",
    "Westford": "Boston", "Wilmington MA": "Boston",

    # New York
    "New York": "New York", "Manhattan": "New York", "Brooklyn": "New York",
    "Queens": "New York", "Bronx": "New York", "Staten Island": "New York",
    "Hoboken": "New York", "Jersey City": "New York", "Newark": "New York",
    "Stamford": "New York", "White Plains": "New York", "Yonkers": "New York",

    # Orange County
    "Irvine": "Orange County", "Anaheim": "Orange County",
    "Santa Ana": "Orange County", "Huntington Beach": "Orange County",
    "Costa Mesa": "Orange County", "Garden Grove": "Orange County",
    "Fullerton": "Orange County", "Orange CA": "Orange County",
    "Orange": "Orange County", "Tustin": "Orange County",
    "Aliso Viejo": "Orange County", "Lake Forest CA": "Orange County",
    "Lake Forest": "Orange County", "Mission Viejo": "Orange County",
    "Newport Beach": "Orange County", "Brea": "Orange County",
    "Buena Park": "Orange County", "Cypress CA": "Orange County",
    "Dana Point": "Orange County", "Fountain Valley": "Orange County",
    "Laguna Hills": "Orange County", "Laguna Niguel": "Orange County",
    "Los Alamitos": "Orange County", "Placentia": "Orange County",
    "Rancho Santa Margarita": "Orange County", "San Clemente": "Orange County",
    "Seal Beach": "Orange County", "Stanton": "Orange County",
    "Westminster CA": "Orange County", "Yorba Linda": "Orange County",

    # Charleston
    "Charleston": "Charleston", "Mount Pleasant SC": "Charleston",
    "Mount Pleasant": "Charleston", "North Charleston": "Charleston",
    "Summerville SC": "Charleston", "Summerville": "Charleston",
    "Goose Creek": "Charleston", "Hanahan": "Charleston",
    "Johns Island": "Charleston", "Ladson": "Charleston",

    # Standalone metros
    "Jacksonville": "Jacksonville",
    "Philadelphia": "Philadelphia", "Wilmington DE": "Philadelphia",
    "Cherry Hill NJ": "Philadelphia", "King of Prussia": "Philadelphia",
    "Norristown": "Philadelphia",
    "Las Vegas": "Las Vegas", "Henderson NV": "Las Vegas",
    "North Las Vegas": "Las Vegas",
    "Portland": "Portland", "Beaverton": "Portland", "Gresham": "Portland",
    "Hillsboro": "Portland", "Lake Oswego": "Portland",
    "Milwaukie": "Portland", "Tigard": "Portland", "Tualatin": "Portland",
    "Vancouver WA": "Portland",
    "San Antonio": "San Antonio", "New Braunfels": "San Antonio",
    "Salt Lake City": "Salt Lake City", "Provo": "Salt Lake City",
    "Ogden": "Salt Lake City", "Sandy UT": "Salt Lake City",
    "Orem UT": "Salt Lake City", "West Valley City": "Salt Lake City",
    "Layton UT": "Salt Lake City",
    "Inland Empire": "Inland Empire", "Riverside CA": "Inland Empire",
    "San Bernardino": "Inland Empire", "Ontario CA": "Inland Empire",
    "Rancho Cucamonga": "Inland Empire", "Fontana": "Inland Empire",
    "Moreno Valley": "Inland Empire",
    "Memphis": "Memphis",
    "Richmond": "Richmond", "Richmond VA": "Richmond",
    "Henrico": "Richmond", "Chesterfield VA": "Richmond",
    "Savannah": "Savannah",
    "Birmingham": "Birmingham", "Hoover AL": "Birmingham",
    "Huntsville": "Huntsville", "Madison AL": "Huntsville",
    "Greenville SC": "Greenville", "Greenville": "Greenville",
    "Spartanburg": "Greenville", "Mauldin": "Greenville",
    "Simpsonville": "Greenville",
    "Gulf Shores": "Gulf Shores", "Pensacola": "Gulf Shores",
    "Mobile AL": "Gulf Shores",
    "Tallahassee": "Tallahassee",
    "Panama City Beach": "Panama City Beach", "Panama City FL": "Panama City Beach",
    "Kansas City": "Kansas City", "Overland Park": "Kansas City",
    "Lenexa": "Kansas City",
    "Charlottesville": "Charlottesville",
    "Hampton Roads": "Hampton Roads", "Norfolk": "Hampton Roads",
    "Norfolk VA": "Hampton Roads", "Virginia Beach": "Hampton Roads",
    "Chesapeake VA": "Hampton Roads", "Suffolk VA": "Hampton Roads",
    "Salinas": "Salinas/Monterey", "Monterey CA": "Salinas/Monterey",
    "Santa Cruz": "Salinas/Monterey",
    "Louisville": "Louisville",
    "Santa Barbara": "Santa Barbara", "Goleta": "Santa Barbara",
    "Ventura": "Santa Barbara", "Oxnard": "Santa Barbara",
    "Camarillo": "Santa Barbara",

    # Baltimore (new macro-market)
    "Baltimore": "Baltimore", "Linthicum Heights": "Baltimore",
    "Annapolis": "Baltimore", "Hunt Valley": "Baltimore", "Laurel": "Baltimore",
    "Columbia MD": "Baltimore", "Wheaton": "Baltimore", "Towson": "Baltimore",

    # Chicago (new macro-market)
    "Chicago": "Chicago",

    # Minneapolis (new macro-market)
    "Minneapolis": "Minneapolis",

    # Chattanooga (new macro-market)
    "Chattanooga": "Chattanooga",

    # Gainesville (new macro-market)
    "Gainesville": "Gainesville",

    # Orange County — self-mapping for when the literal label appears
    "Orange County": "Orange County",

    # South Florida / Bay Area generic labels
    "South Florida": "Miami/Fort Lauderdale",
    "Bay Area": "San Francisco-East Bay",
    "San Francisco Bay Area": "San Francisco-East Bay",
    "Southern California": "Los Angeles",

    # Additional Miami/Fort Lauderdale
    "Doral": "Miami/Fort Lauderdale", "Coconut Creek": "Miami/Fort Lauderdale",

    # Additional Seattle
    "Newcastle": "Seattle", "Woodinville": "Seattle", "Mill Creek": "Seattle",
    "Mercer Island": "Seattle",

    # Additional San Francisco-East Bay
    "Dublin": "San Francisco-East Bay", "Emeryville": "San Francisco-East Bay",
    "Union City": "San Francisco-East Bay", "San Rafael": "San Francisco-East Bay",
    "Tiburon": "San Francisco-East Bay",

    # Additional San Jose
    "Menlo Park": "San Jose",

    # Additional Los Angeles
    "Studio City": "Los Angeles", "Calabasas": "Los Angeles",
    "Agoura Hills": "Los Angeles", "Artesia": "Los Angeles",
    "Playa Vista": "Los Angeles", "Valencia": "Los Angeles",
    "Valley Village": "Los Angeles", "Newbury Park": "Los Angeles",
    "La Habra": "Los Angeles", "Walnut": "Los Angeles",
    "Chino Hills": "Los Angeles",

    # Additional San Diego
    "Bonita": "San Diego", "Spring Valley": "San Diego",

    # Additional New York (northern NJ + Long Island)
    "Roseland": "New York", "West Windsor": "New York",
    "Florham Park": "New York", "Wharton": "New York",
    "Pine Brook": "New York", "Old Bridge": "New York",
    "Boonton": "New York", "Piscataway": "New York",
    "Maplewood": "New York", "Bloomfield": "New York",
    "Union NJ": "New York", "North Bergen": "New York",
    "Bloomingdale": "New York", "Teaneck": "New York",
    "Long Island City": "New York", "Amityville": "New York",
    "Westbury": "New York", "Garden City": "New York",
    "Rockville Centre": "New York", "Smithtown": "New York",
    "Melville": "New York", "Island Park": "New York",
    "Great Neck": "New York", "Huntington Station": "New York",
    "Baldwin Place": "New York", "Harrison": "New York",

    # Additional Boston
    "South Easton": "Boston", "Northborough": "Boston",
    "Chestnut Hill": "Boston", "Saugus": "Boston", "Hingham": "Boston",
    "Plymouth": "Boston",

    # Additional Washington, DC
    "North Potomac": "Washington, DC", "North Bethesda": "Washington, DC",
    "Vienna": "Washington, DC", "Tysons Corner": "Washington, DC",

    # Additional Dallas/Fort Worth
    "Benbrook": "Dallas/Fort Worth",

    # Additional Austin
    "Bee Cave": "Austin",

    # Additional Denver
    "Castle Rock": "Denver", "Lafayette": "Denver",

    # Additional Charlotte
    "Mooresville NC": "Charlotte",

    # APT 1, KENNESAW → Atlanta (bad data)
    "APT 1, KENNESAW": "Atlanta",
}

# Secondary map for "City, ST" format entries that can't be handled by
# stripping the state suffix alone (ambiguous city names, abbreviations, etc.)
_CITY_ST_MAP = {
    # Baltimore
    "Linthicum Heights, MD": "Baltimore", "Annapolis, MD": "Baltimore",
    "Hunt Valley, MD": "Baltimore", "Laurel, MD": "Baltimore",
    "Columbia, MD": "Baltimore", "Wheaton, MD": "Baltimore", "Towson, MD": "Baltimore",

    # Miami/Fort Lauderdale
    "South Miami, FL": "Miami/Fort Lauderdale", "Miramar, FL": "Miami/Fort Lauderdale",
    "Doral, FL": "Miami/Fort Lauderdale", "Coconut Creek, FL": "Miami/Fort Lauderdale",
    "West Palm Beach, FL": "Miami/Fort Lauderdale", "Hialeah, FL": "Miami/Fort Lauderdale",
    "Margate, FL": "Miami/Fort Lauderdale",

    # San Francisco-East Bay
    "Dublin, CA": "San Francisco-East Bay", "Emeryville, CA": "San Francisco-East Bay",
    "Union City, CA": "San Francisco-East Bay",

    # San Francisco
    "San Bruno, CA": "San Francisco", "Pacifica, CA": "San Francisco",

    # Charlotte
    "Mooresville, NC": "Charlotte",

    # Northern NJ → New York
    "Roseland, NJ": "New York", "West Windsor, NJ": "New York",
    "Somerville, NJ": "New York", "Florham Park, NJ": "New York",
    "Wharton, NJ": "New York", "Pine Brook, NJ": "New York",
    "Old Bridge, NJ": "New York", "Boonton, NJ": "New York",
    "Piscataway, NJ": "New York", "Maplewood, NJ": "New York",
    "Bloomfield, NJ": "New York", "Union, NJ": "New York",
    "North Bergen, NJ": "New York", "Bloomingdale, NJ": "New York",
    "Teaneck, NJ": "New York",

    # Long Island / NY suburbs → New York
    "Long Island City, NY": "New York", "Amityville, NY": "New York",
    "Westbury, NY": "New York", "Garden City, NY": "New York",
    "Rockville Centre, NY": "New York", "Smithtown, NY": "New York",
    "Melville, NY": "New York", "Island Park, NY": "New York",
    "Great Neck, NY": "New York", "Huntington Station, NY": "New York",
    "Baldwin Place, NY": "New York", "Harrison, NY": "New York",

    # Austin
    "Pflugerville, TX": "Austin", "Georgetown, TX": "Austin", "Bee Cave, TX": "Austin",

    # Los Angeles
    "Woodland Hills, CA": "Los Angeles", "Studio City, CA": "Los Angeles",
    "Calabasas, CA": "Los Angeles", "Agoura Hills, CA": "Los Angeles",
    "Pasadena, CA": "Los Angeles", "Simi Valley, CA": "Los Angeles",
    "Glendora, CA": "Los Angeles", "Pomona, CA": "Los Angeles",
    "San Dimas, CA": "Los Angeles", "Artesia, CA": "Los Angeles",
    "Monrovia, CA": "Los Angeles", "Canoga Park, CA": "Los Angeles",
    "Encino, CA": "Los Angeles", "Cerritos, CA": "Los Angeles",
    "Chino Hills, CA": "Los Angeles", "Thousand Oaks, CA": "Los Angeles",
    "Santa Monica, CA": "Los Angeles",

    # Santa Barbara
    "Camarillo, CA": "Santa Barbara",

    # Orange County
    "Seal Beach, CA": "Orange County", "Huntington Beach, CA": "Orange County",
    "Brea, CA": "Orange County", "Rancho Santa Marg, CA": "Orange County",
    "Lake Forest, CA": "Orange County", "Mission Viejo, CA": "Orange County",

    # San Diego
    "Vista, CA": "San Diego", "La Mesa, CA": "San Diego",
    "San Marcos, CA": "San Diego",

    # Dallas/Fort Worth
    "Carrollton, TX": "Dallas/Fort Worth", "Lewisville, TX": "Dallas/Fort Worth",
    "Flower Mound, TX": "Dallas/Fort Worth", "Benbrook, TX": "Dallas/Fort Worth",
    "Allen, TX": "Dallas/Fort Worth", "Addison, TX": "Dallas/Fort Worth",

    # Boston
    "Somerville, MA": "Boston", "Acton, MA": "Boston",
    "South Easton, MA": "Boston", "Northborough, MA": "Boston",
    "Chestnut Hill, MA": "Boston", "Saugus, MA": "Boston",
    "Natick, MA": "Boston", "Hingham, MA": "Boston",
    "North Andover, MA": "Boston", "Norwood, MA": "Boston",
    "Framingham, MA": "Boston", "Sudbury, MA": "Boston",
    "Plymouth, MA": "Boston", "Bedford, MA": "Boston", "Milford, MA": "Boston",

    # Washington, DC
    "North Potomac, MD": "Washington, DC", "North Bethesda, MD": "Washington, DC",
    "Vienna, VA": "Washington, DC", "Tysons Corner, VA": "Washington, DC",

    # Denver
    "Littleton, CO": "Denver", "Castle Rock, CO": "Denver", "Lafayette, CO": "Denver",

    # Seattle
    "Lynnwood, WA": "Seattle", "Newcastle, WA": "Seattle",
    "Woodinville, WA": "Seattle",

    # San Jose
    "Menlo Park, CA": "San Jose",

    # Atlanta
    "Kennesaw, GA": "Atlanta",
}

# ─────────────────────────────────────────────────────────────────────────────
# STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
GS_NAVY   = "003B70"
WHITE     = "FFFFFF"
ALT_ROW   = "F5F5F5"
LT_GREEN  = "C6EFCE"
LT_RED    = "FFC7CE"
THIN_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

HEADER_FONT   = Font(name="Arial", bold=True, color=WHITE, size=10)
HEADER_FILL   = PatternFill("solid", fgColor=GS_NAVY)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT     = Font(name="Arial", size=9)
BODY_ALIGN    = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
TITLE_FONT    = Font(name="Arial", bold=True, size=12, color=GS_NAVY)
SUBHEAD_FONT  = Font(name="Arial", bold=True, size=10, color=GS_NAVY)
NUM_CURRENCY  = '$#,##0'
NUM_PCT       = '0.0%'
NUM_COMMA     = '#,##0'

# ─────────────────────────────────────────────────────────────────────────────
# FIXED COLOR PALETTES — consistent across all charts
# ─────────────────────────────────────────────────────────────────────────────

# REIT colors (used in any multi-REIT chart)
REIT_COLORS = {
    "MAA":  "2E8B57",  # sea green
    "CPT":  "1F77B4",  # steel blue
    "EQR":  "D62728",  # crimson
    "AVB":  "FF7F0E",  # orange
    "UDR":  "9467BD",  # purple
    "ESS":  "E377C2",  # pink
    "INVH": "8C564B",  # brown
    "AMH":  "17BECF",  # teal
}

# Market colors (used in any multi-market chart)
MARKET_COLORS = {
    "Atlanta":              "D62728",  # red
    "Dallas/Fort Worth":    "1F77B4",  # blue
    "Houston":              "FF7F0E",  # orange
    "Phoenix":              "2CA02C",  # green
    "Tampa":                "9467BD",  # purple
    "Orlando":              "8C564B",  # brown
    "Charlotte":            "E377C2",  # pink
    "Nashville":            "7F7F7F",  # gray
    "Raleigh/Durham":       "BCBD22",  # olive
    "Austin":               "17BECF",  # teal
    "Denver":               "AEC7E8",  # light blue
    "Miami/Fort Lauderdale":"FF9896",  # salmon
    "Washington, DC":       "003B70",  # navy
    "New York":             "393B79",  # dark blue
    "Boston":               "637939",  # dark olive
    "San Francisco":        "8C6D31",  # dark gold
    "San Jose":             "843C39",  # dark red
    "Los Angeles":          "7B4173",  # plum
    "San Diego":            "5254A3",  # indigo
    "Seattle":              "6B6ECF",  # periwinkle
    "Orange County":        "E7969C",  # light rose
    "San Francisco-East Bay":"B5CF6B", # lime
    "Jacksonville":         "CE6DBD",  # orchid
    "Charleston":           "DE9ED6",  # lavender
    "Savannah":             "D6616B",  # dark salmon
    "Las Vegas":            "E7BA52",  # gold
    "Portland":             "3182BD",  # medium blue
    "Salt Lake City":       "6BAED6",  # sky blue
    "Inland Empire":        "FD8D3C",  # dark orange
    "San Antonio":          "FDAE6B",  # light orange
    "Philadelphia":         "31A354",  # forest green
    "Baltimore":            "74C476",  # medium green
    "Richmond":             "756BB1",  # medium purple
    "Greenville":           "9E9AC8",  # light purple
    "Memphis":              "636363",  # dark gray
    "Birmingham":           "969696",  # medium gray
    "Huntsville":           "BDBDBD",  # light gray
    "Chicago":              "E6550D",  # burnt orange
    "Minneapolis":          "A1D99B",  # mint
    "Kansas City":          "C7E9C0",  # pale green
    "Hampton Roads":        "DADAEB",  # pale lavender
    "Chattanooga":          "9C9EDE",  # blue-gray
    "Gainesville":          "CEDB9C",  # sage
    "Tallahassee":          "E7CB94",  # tan
    "Gulf Shores":          "F7B6D2",  # light pink
    "Panama City Beach":    "FDD0A2",  # peach
    "Charlottesville":      "C49C94",  # dusty rose
    "Louisville":           "C5B0D5",  # thistle
    "Colorado Springs":     "98DF8A",  # light green
    "Santa Barbara":        "DBDB8D",  # khaki
    "Salinas/Monterey":     "ADB5BD",  # blue-gray
}

# Fallback palette for anything not in the dicts above
_FALLBACK_COLORS = [
    "1F77B4", "FF7F0E", "2CA02C", "D62728", "9467BD",
    "8C564B", "E377C2", "7F7F7F", "BCBD22", "17BECF",
    "AEC7E8", "FFBB78", "98DF8A", "FF9896", "C5B0D5",
]


def _apply_series_colors(chart, series_names, color_dict):
    """Apply consistent colors to chart series based on a name->color dict."""
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    for i, series in enumerate(chart.series):
        name = series_names[i] if i < len(series_names) else None
        if name and name in color_dict:
            hex_color = color_dict[name]
        else:
            hex_color = _FALLBACK_COLORS[i % len(_FALLBACK_COLORS)]
        series.graphicalProperties.line.solidFill = hex_color
        # Also set marker color for line charts
        if hasattr(series, 'marker') and series.marker:
            series.marker.graphicalProperties.solidFill = hex_color


def _apply_bar_colors(chart, series_names, color_dict):
    """Apply consistent colors to bar chart series."""
    for i, series in enumerate(chart.series):
        name = series_names[i] if i < len(series_names) else None
        if name and name in color_dict:
            hex_color = color_dict[name]
        else:
            hex_color = _FALLBACK_COLORS[i % len(_FALLBACK_COLORS)]
        series.graphicalProperties.solidFill = hex_color


# ─────────────────────────────────────────────────────────────────────────────
# FILENAME PARSING (must be before download functions that use it)
# ─────────────────────────────────────────────────────────────────────────────

# Patterns: {ticker}_raw_{YYYY-MM-DD}.csv
#           {ticker}_raw_{YYYY-MM-DD}_part1.csv / _part2.csv
FNAME_RE = re.compile(r'^(.+?)_raw_(\d{4}-\d{2}-\d{2})(?:_part(\d+))?\.csv$', re.IGNORECASE)


def parse_filename(fname):
    """Returns (ticker, scrape_date, part_num_or_None)."""
    m = FNAME_RE.match(fname)
    if not m:
        return None, None, None
    ticker     = m.group(1).upper()
    date_str   = m.group(2)
    part_num   = int(m.group(3)) if m.group(3) else None
    return ticker, date_str, part_num


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 & 2: GITHUB DOWNLOAD + LOCAL CACHE
# ─────────────────────────────────────────────────────────────────────────────

def _api_headers():
    h = {"Accept": "application/vnd.github+json"}
    if GITHUB_TOKEN:
        h["Authorization"] = f"token {GITHUB_TOKEN}"
    return h


def _load_manifest():
    manifest_path = os.path.join(CACHE_DIR, "cache_manifest.json")
    if os.path.exists(manifest_path):
        with open(manifest_path, "r") as f:
            return json.load(f)
    return {}


def _save_manifest(manifest):
    os.makedirs(CACHE_DIR, exist_ok=True)
    manifest_path = os.path.join(CACHE_DIR, "cache_manifest.json")
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2)


def _cached_path(filename):
    return os.path.join(CACHE_DIR, filename)


def list_github_files(path):
    """Return list of file metadata dicts from GitHub contents API."""
    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{path}"
    resp = requests.get(url, headers=_api_headers(), timeout=30)
    if resp.status_code == 404:
        print(f"  [WARNING] Path not found in GitHub repo: {path}")
        return []
    resp.raise_for_status()
    return resp.json()


def download_file(download_url):
    """Download raw file bytes from GitHub."""
    resp = requests.get(download_url, headers=_api_headers(), timeout=60)
    resp.raise_for_status()
    return resp.content


def _snap_to_saturday(dt):
    """Snap a date to the Saturday of that week (Saturday=5 in weekday())."""
    offset = (dt.weekday() - 5) % 7  # 0 if already Saturday
    return dt - timedelta(days=offset)


def _parse_date_from_filename(fname):
    """Extract date string from a CSV filename, return as datetime.date or None."""
    m = FNAME_RE.match(fname)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(2), "%Y-%m-%d").date()
    except ValueError:
        return None


def fetch_latest_2_weeks_csvs():
    """
    List all CSVs in DATA_PATH on GitHub, parse dates from filenames,
    identify the 2 most recent Saturday-normalized periods, and only
    download CSVs from those 2 periods. Returns list of (filename, bytes_content).
    """
    os.makedirs(CACHE_DIR, exist_ok=True)
    manifest = _load_manifest()

    print(f"\n[Step 1] Listing files in GitHub repo: {GITHUB_OWNER}/{GITHUB_REPO}/{DATA_PATH}")
    all_items = list_github_files(DATA_PATH)
    csv_items = [f for f in all_items if isinstance(f, dict) and f.get("name", "").endswith(".csv")]
    print(f"  Found {len(csv_items)} CSV file(s) total in repo")

    # Parse dates and snap to Saturday
    file_periods = {}  # fname -> saturday date
    for item in csv_items:
        fname = item["name"]
        raw_date = _parse_date_from_filename(fname)
        if raw_date is None:
            continue
        sat_date = _snap_to_saturday(raw_date)
        file_periods[fname] = sat_date

    # Find the 2 most recent Saturday periods
    unique_periods = sorted(set(file_periods.values()), reverse=True)
    if not unique_periods:
        print("  [ERROR] No dated CSV files found.")
        return []

    target_periods = set(unique_periods[:2])
    period_labels = ", ".join(str(p) for p in sorted(target_periods))
    print(f"  Target periods (latest 2 Saturdays): {period_labels}")

    # Filter to only files in those periods
    target_items = [
        item for item in csv_items
        if item["name"] in file_periods and file_periods[item["name"]] in target_periods
    ]
    print(f"  Downloading {len(target_items)} CSV file(s) from target periods")

    results = []
    downloaded = 0
    cached_hits = 0

    for item in target_items:
        fname = item["name"]
        sha   = item.get("sha", "")
        durl  = item.get("download_url", "")
        cache_file = _cached_path(fname)

        if manifest.get(fname) == sha and os.path.exists(cache_file):
            with open(cache_file, "rb") as fh:
                content = fh.read()
            cached_hits += 1
        else:
            print(f"  Downloading: {fname}")
            content = download_file(durl)
            with open(cache_file, "wb") as fh:
                fh.write(content)
            manifest[fname] = sha
            downloaded += 1

        results.append((fname, content))

    _save_manifest(manifest)
    print(f"  Downloaded: {downloaded} file(s), cache hits: {cached_hits}")
    return results


def fetch_registry():
    """Download unit_registry.csv, with caching."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    manifest = _load_manifest()

    fname = "unit_registry.csv"
    # Get metadata for the registry file
    parent_path = os.path.dirname(REGISTRY_PATH)
    reg_filename = os.path.basename(REGISTRY_PATH)

    try:
        url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{REGISTRY_PATH}"
        resp = requests.get(url, headers=_api_headers(), timeout=30)
        if resp.status_code == 404:
            print(f"  [INFO] Registry file not found ({REGISTRY_PATH}), skipping.")
            return None
        resp.raise_for_status()
        item = resp.json()
        sha  = item.get("sha", "")
        durl = item.get("download_url", "")
        cache_file = _cached_path(fname)

        if manifest.get(fname) == sha and os.path.exists(cache_file):
            with open(cache_file, "rb") as fh:
                content = fh.read()
        else:
            print(f"  Downloading registry: {reg_filename}")
            content = download_file(durl)
            with open(cache_file, "wb") as fh:
                fh.write(content)
            manifest[fname] = sha
            _save_manifest(manifest)

        return pd.read_csv(io.BytesIO(content))
    except Exception as e:
        print(f"  [WARNING] Could not fetch registry: {e}")
        return None


def fetch_summary_history():
    """
    Download summary_history.csv from GitHub (if it exists).
    Returns a DataFrame or None if not found.
    """
    print(f"\n[Step 1c] Fetching summary history from GitHub...")
    try:
        url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{SUMMARY_PATH}"
        resp = requests.get(url, headers=_api_headers(), timeout=30)
        if resp.status_code == 404:
            print(f"  [INFO] summary_history.csv not found yet (first run). Will create.")
            return None
        resp.raise_for_status()
        item = resp.json()
        durl = item.get("download_url", "")
        content = download_file(durl)
        df = pd.read_csv(io.BytesIO(content))
        print(f"  Loaded summary_history.csv: {len(df):,} existing rows")
        return df
    except Exception as e:
        print(f"  [WARNING] Could not fetch summary_history.csv: {e}")
        return None


def save_summary_history(summary_df):
    """Save summary_history.csv locally for git push by weekly_run.py."""
    os.makedirs(LOCAL_SUMMARY_DIR, exist_ok=True)
    out_path = os.path.join(LOCAL_SUMMARY_DIR, "summary_history.csv")
    summary_df.to_csv(out_path, index=False)
    print(f"  Saved summary_history.csv: {len(summary_df):,} rows -> {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2b: MERGE PART FILES
# ─────────────────────────────────────────────────────────────────────────────

def merge_parts(raw_files):
    """
    Detect paired _part1/_part2 files and concatenate them.
    Returns list of (base_filename, bytes_content) with parts merged.
    """
    # Group by (ticker, date) key
    groups = defaultdict(dict)
    singles = []

    for fname, content in raw_files:
        ticker, date_str, part_num = parse_filename(fname)
        if ticker is None:
            singles.append((fname, content))
            continue
        if part_num is not None:
            groups[(ticker, date_str)][part_num] = (fname, content)
        else:
            singles.append((fname, content))

    merged = list(singles)
    for (ticker, date_str), parts in groups.items():
        sorted_parts = sorted(parts.items())  # sort by part number
        combined_bytes = b"\n".join(
            content if i == 0 else _strip_header(content)
            for i, (_, (_, content)) in enumerate(sorted_parts)
        )
        base_fname = f"{ticker}_raw_{date_str}.csv"
        merged.append((base_fname, combined_bytes))
        print(f"  Merged {len(sorted_parts)} parts -> {base_fname}")

    return merged


def _strip_header(content_bytes):
    """Remove the first (header) line from CSV bytes."""
    lines = content_bytes.split(b"\n")
    return b"\n".join(lines[1:]) if len(lines) > 1 else b""


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: COMBINE INTO PANEL
# ─────────────────────────────────────────────────────────────────────────────

EXPECTED_COLS = [
    "scrape_date", "reit", "market", "community", "unit_id", "beds", "sqft",
    "rent", "has_concession", "concession_hardness", "concession_raw",
    "concession_type", "concession_value", "effective_monthly_rent"
]


def build_panel(file_list):
    """
    Parse all CSVs, add reit/scrape_date columns, combine into a panel DataFrame.
    """
    dfs = []
    for fname, content in file_list:
        ticker, date_str, _ = parse_filename(fname)
        if ticker is None:
            print(f"  [SKIP] Unrecognised filename: {fname}")
            continue
        try:
            df = pd.read_csv(io.BytesIO(content), low_memory=False)
        except Exception as e:
            print(f"  [ERROR] Could not parse {fname}: {e}")
            continue

        df["reit"]        = ticker
        df["scrape_date"] = date_str

        dfs.append(df)

    if not dfs:
        raise RuntimeError("No valid CSV files found. Check GitHub token and DATA_PATH.")

    panel = pd.concat(dfs, ignore_index=True)

    # Ensure all expected columns exist (fill with NaN if absent)
    for col in EXPECTED_COLS:
        if col not in panel.columns:
            panel[col] = float("nan")

    # Normalise beds column
    bed_map = {"Studio": 0, "studio": 0, "STUDIO": 0}
    panel["beds"] = panel["beds"].apply(
        lambda x: bed_map.get(str(x), x) if pd.notna(x) else x
    )
    panel["beds"] = pd.to_numeric(panel["beds"], errors="coerce")

    # Deduplicate
    before = len(panel)
    panel = panel.drop_duplicates(subset=["unit_id", "scrape_date"], keep="last")
    after = len(panel)
    if before != after:
        print(f"  Deduplication: {before:,} -> {after:,} rows (removed {before - after:,} dupes)")

    # Sort
    panel["scrape_date"] = pd.to_datetime(panel["scrape_date"], errors="coerce")

    # ── Normalize scrape dates to weekly periods ──
    # Scrapes run over a weekend (e.g. Sat night → Sun morning). Snap each date
    # to the Saturday of that week so Mar-30 and Mar-31 become the same period,
    # and Apr-4 and Apr-5 become the same period, etc.
    # Formula: date - (weekday offset from Saturday).  Saturday=5 in pandas.
    def snap_to_saturday(dt):
        if pd.isna(dt):
            return dt
        offset = (dt.weekday() - 5) % 7  # 0 if already Saturday
        return dt - pd.Timedelta(days=offset)

    panel["scrape_date"] = panel["scrape_date"].apply(snap_to_saturday)

    # Re-deduplicate after date normalization (same unit may appear on both
    # Mar-30 and Mar-31 → now both are Mar-29 Saturday; keep the latest)
    before_dedup2 = len(panel)
    panel = panel.drop_duplicates(subset=["unit_id", "scrape_date"], keep="last")
    after_dedup2 = len(panel)
    if before_dedup2 != after_dedup2:
        print(f"  Post-normalization dedup: {before_dedup2:,} -> {after_dedup2:,} rows "
              f"(removed {before_dedup2 - after_dedup2:,} dupes)")

    panel = panel.sort_values(["scrape_date", "reit", "unit_id"]).reset_index(drop=True)

    n_periods = panel["scrape_date"].nunique()
    period_labels = ", ".join(str(d.date()) for d in sorted(panel["scrape_date"].unique()))
    print(f"  Panel: {len(panel):,} rows, {n_periods} period(s) [{period_labels}], "
          f"{panel['reit'].nunique()} REIT(s)")

    # ── ESS concession fix: week 1 data was scraped with a bug that missed
    # the property-offer-cta DOM element. Null out concession fields for ESS
    # on the earliest period so they're excluded from all comparisons.
    # Rent data is unaffected and remains valid.
    ess_first_mask = (panel["reit"] == "ESS") & (panel["scrape_date"] == panel["scrape_date"].min())
    if ess_first_mask.any():
        # Convert has_concession to nullable before assigning NaN
        if "has_concession" in panel.columns:
            panel["has_concession"] = panel["has_concession"].astype("object")
        conc_cols = ["has_concession", "concession_hardness", "concession_raw",
                     "concession_type", "concession_value", "concession_pct_lease_value",
                     "concession_pct_lease_term", "effective_monthly_rent"]
        for col in conc_cols:
            if col in panel.columns:
                panel.loc[ess_first_mask, col] = None
        print(f"  [FIX] Nulled ESS concession fields for first period ({panel['scrape_date'].min().date()}) "
              f"— {ess_first_mask.sum():,} rows. Scraper bug (fixed week 2).")

    return panel


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: MACRO-MARKET MAPPING
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_macro_market(market_value):
    """
    Resolve a market string to its macro-market using a multi-step fallback:
      1. Direct lookup in MACRO_MAP
      2. Direct lookup in _CITY_ST_MAP (handles "City, ST" format)
      3. Strip ", XX" state suffix and retry MACRO_MAP  (e.g. "Addison, TX" → "Addison")
      4. Strip ", XX" and retry with "City ST" format   (e.g. "Addison TX")
      5. Return "Other"
    """
    if not isinstance(market_value, str) or not market_value.strip():
        return "Other"

    val = market_value.strip()

    # 1. Direct lookup
    if val in MACRO_MAP:
        return MACRO_MAP[val]

    # 2. "City, ST" secondary map
    if val in _CITY_ST_MAP:
        return _CITY_ST_MAP[val]

    # 3-4. Try stripping state suffix  ("City, ST" → "City" and "City ST")
    m = re.match(r'^(.+),\s*([A-Z]{2})$', val)
    if m:
        city, state = m.group(1).strip(), m.group(2)
        city_st = f"{city} {state}"
        if city_st in MACRO_MAP:
            return MACRO_MAP[city_st]
        if city in MACRO_MAP:
            return MACRO_MAP[city]

    return "Other"


def apply_macro_map(df):
    if "market" not in df.columns:
        df["macro_market"] = "Other"
        return df

    df["macro_market"] = df["market"].apply(_resolve_macro_market)

    # Report unmapped markets
    other_mask = df["macro_market"] == "Other"
    if other_mask.any():
        unmapped = df.loc[other_mask, "market"].value_counts()
        print(f"\n  [INFO] {len(unmapped)} market value(s) mapped to 'Other':")
        for mkt, cnt in unmapped.items():
            print(f"    {mkt!r}: {cnt:,} rows")
    else:
        print("  All market values mapped successfully.")

    return df


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: SAME-PROPERTY ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

def _safe_div(a, b):
    """Safe division returning None if b is 0 or either is NaN."""
    if pd.isna(a) or pd.isna(b) or b == 0:
        return None
    return a / b


def compute_same_property(df):
    """
    For each consecutive pair of scrape_dates, compute same-property metrics
    at REIT × macro_market × beds granularity.
    Returns sp_df or empty DataFrame.
    """
    dates = sorted(df["scrape_date"].dropna().unique())
    if len(dates) < 2:
        print("  [INFO] Only 1 scrape date — same-property analysis requires Week 2+")
        return pd.DataFrame()

    records = []
    for i in range(1, len(dates)):
        prev_date = dates[i - 1]
        curr_date = dates[i]
        prev = df[df["scrape_date"] == prev_date].copy()
        curr = df[df["scrape_date"] == curr_date].copy()

        same_prop_ids = set(prev["unit_id"].dropna()) & set(curr["unit_id"].dropna())
        if not same_prop_ids:
            continue

        prev_sp = prev[prev["unit_id"].isin(same_prop_ids)].copy()
        curr_sp = curr[curr["unit_id"].isin(same_prop_ids)].copy()

        # Compute unit-level rent PSF and effective rent PSF
        for frame in [prev_sp, curr_sp]:
            frame["_rent_psf"] = frame.apply(
                lambda r: r["rent"] / r["sqft"] if pd.notna(r["sqft"]) and r["sqft"] > 0 and pd.notna(r["rent"]) else None,
                axis=1,
            )
            frame["_eff_rent_psf"] = frame.apply(
                lambda r: r["effective_monthly_rent"] / r["sqft"]
                if pd.notna(r.get("effective_monthly_rent")) and pd.notna(r["sqft"]) and r["sqft"] > 0
                else None,
                axis=1,
            )

        # Group by REIT x macro_market x beds
        key_cols = ["reit", "macro_market", "beds"]

        prev_grp = prev_sp.groupby(key_cols, dropna=False).agg(
            sp_avg_rent_prev=("rent", "mean"),
            sp_concession_rate_prev=("has_concession", "mean"),
            sp_count_prev=("unit_id", "count"),
            sp_avg_rent_psf_prev=("_rent_psf", "mean"),
            sp_avg_eff_rent_prev=("effective_monthly_rent", "mean"),
            sp_avg_eff_rent_psf_prev=("_eff_rent_psf", "mean"),
        ).reset_index()

        curr_grp = curr_sp.groupby(key_cols, dropna=False).agg(
            sp_avg_rent_curr=("rent", "mean"),
            sp_concession_rate_curr=("has_concession", "mean"),
            sp_count_curr=("unit_id", "count"),
            sp_avg_rent_psf_curr=("_rent_psf", "mean"),
            sp_avg_eff_rent_curr=("effective_monthly_rent", "mean"),
            sp_avg_eff_rent_psf_curr=("_eff_rent_psf", "mean"),
        ).reset_index()

        merged = pd.merge(prev_grp, curr_grp, on=key_cols, how="inner")
        merged["date_curr"]  = curr_date
        merged["date_prev"]  = prev_date
        merged["sp_count"]   = merged["sp_count_curr"]
        merged["sp_wow_pct"] = (
            (merged["sp_avg_rent_curr"] - merged["sp_avg_rent_prev"])
            / merged["sp_avg_rent_prev"]
        )
        merged["sp_wow_pct_psf"] = merged.apply(
            lambda r: _safe_div(r["sp_avg_rent_psf_curr"] - r["sp_avg_rent_psf_prev"], r["sp_avg_rent_psf_prev"]),
            axis=1,
        )
        merged["sp_wow_pct_eff"] = merged.apply(
            lambda r: _safe_div(r["sp_avg_eff_rent_curr"] - r["sp_avg_eff_rent_prev"], r["sp_avg_eff_rent_prev"]),
            axis=1,
        )
        merged["sp_wow_pct_eff_psf"] = merged.apply(
            lambda r: _safe_div(r["sp_avg_eff_rent_psf_curr"] - r["sp_avg_eff_rent_psf_prev"], r["sp_avg_eff_rent_psf_prev"]),
            axis=1,
        )

        records.append(merged)

    if not records:
        return pd.DataFrame()

    sp_df = pd.concat(records, ignore_index=True)
    out_cols = [
        "date_curr", "date_prev", "reit", "macro_market", "beds",
        "sp_count", "sp_avg_rent_curr", "sp_avg_rent_prev", "sp_wow_pct",
        "sp_concession_rate_curr", "sp_concession_rate_prev",
        "sp_avg_rent_psf_curr", "sp_avg_rent_psf_prev", "sp_wow_pct_psf",
        "sp_avg_eff_rent_curr", "sp_avg_eff_rent_prev", "sp_wow_pct_eff",
        "sp_avg_eff_rent_psf_curr", "sp_avg_eff_rent_psf_prev", "sp_wow_pct_eff_psf",
    ]
    for c in out_cols:
        if c not in sp_df.columns:
            sp_df[c] = float("nan")

    print(f"  Same-property analysis: {len(sp_df):,} rows across {len(dates) - 1} period(s)")
    return sp_df[out_cols]


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5b: SUMMARY HISTORY — persistent accumulator
# ─────────────────────────────────────────────────────────────────────────────

SUMMARY_COLS = [
    "scrape_date", "reit", "macro_market", "beds",
    "listing_count", "avg_rent", "median_rent", "avg_sqft",
    "rent_per_sqft", "concession_rate", "avg_concession_value",
    "avg_rent_psf", "median_rent_psf", "avg_eff_rent", "avg_eff_rent_psf",
    "sp_count", "sp_avg_rent_curr", "sp_avg_rent_prev",
    "sp_wow_pct", "sp_concession_rate_curr", "sp_concession_rate_prev",
    "sp_avg_rent_psf_curr", "sp_avg_rent_psf_prev", "sp_wow_pct_psf",
    "sp_avg_eff_rent_curr", "sp_avg_eff_rent_prev", "sp_wow_pct_eff",
    "sp_avg_eff_rent_psf_curr", "sp_avg_eff_rent_psf_prev", "sp_wow_pct_eff_psf",
]


def build_current_summary(df, sp_df):
    """
    Build summary rows for the latest scrape date in df.
    Returns DataFrame with columns matching SUMMARY_COLS.
    """
    latest_date = df["scrape_date"].max()
    df_latest = df[df["scrape_date"] == latest_date].copy()

    grp = (
        df_latest.groupby(["reit", "macro_market", "beds"], dropna=False)
        .agg(
            listing_count=("unit_id", "count"),
            avg_rent=("rent", "mean"),
            median_rent=("rent", "median"),
            avg_sqft=("sqft", "mean"),
            concession_rate=("has_concession", "mean"),
            avg_concession_value=("concession_value", "mean"),
        )
        .reset_index()
    )

    # Compute rent_per_sqft
    grp["rent_per_sqft"] = grp.apply(
        lambda r: r["avg_rent"] / r["avg_sqft"] if pd.notna(r["avg_sqft"]) and r["avg_sqft"] > 0 else None,
        axis=1,
    )

    # Compute rent PSF metrics (unit-level avg, not avg/avg)
    df_latest["_rent_psf"] = df_latest.apply(
        lambda r: r["rent"] / r["sqft"] if pd.notna(r["sqft"]) and r["sqft"] > 0 and pd.notna(r["rent"]) else None,
        axis=1,
    )
    psf_grp = (
        df_latest[df_latest["_rent_psf"].notna()]
        .groupby(["reit", "macro_market", "beds"], dropna=False)
        .agg(avg_rent_psf=("_rent_psf", "mean"), median_rent_psf=("_rent_psf", "median"))
        .reset_index()
    )
    grp = pd.merge(grp, psf_grp, on=["reit", "macro_market", "beds"], how="left")

    # Compute effective rent metrics
    df_latest["_eff_rent_psf"] = df_latest.apply(
        lambda r: r["effective_monthly_rent"] / r["sqft"]
        if pd.notna(r.get("effective_monthly_rent")) and pd.notna(r["sqft"]) and r["sqft"] > 0
        else None,
        axis=1,
    )
    eff_grp = (
        df_latest[df_latest["effective_monthly_rent"].notna()]
        .groupby(["reit", "macro_market", "beds"], dropna=False)
        .agg(avg_eff_rent=("effective_monthly_rent", "mean"))
        .reset_index()
    )
    grp = pd.merge(grp, eff_grp, on=["reit", "macro_market", "beds"], how="left")

    eff_psf_grp = (
        df_latest[df_latest["_eff_rent_psf"].notna()]
        .groupby(["reit", "macro_market", "beds"], dropna=False)
        .agg(avg_eff_rent_psf=("_eff_rent_psf", "mean"))
        .reset_index()
    )
    grp = pd.merge(grp, eff_psf_grp, on=["reit", "macro_market", "beds"], how="left")

    grp["scrape_date"] = str(latest_date)[:10]

    # Merge same-property columns if available
    sp_cols = ["sp_count", "sp_avg_rent_curr", "sp_avg_rent_prev",
               "sp_wow_pct", "sp_concession_rate_curr", "sp_concession_rate_prev",
               "sp_avg_rent_psf_curr", "sp_avg_rent_psf_prev", "sp_wow_pct_psf",
               "sp_avg_eff_rent_curr", "sp_avg_eff_rent_prev", "sp_wow_pct_eff",
               "sp_avg_eff_rent_psf_curr", "sp_avg_eff_rent_psf_prev", "sp_wow_pct_eff_psf"]
    for c in sp_cols:
        grp[c] = None

    if not sp_df.empty:
        sp_latest = sp_df[sp_df["date_curr"] == sp_df["date_curr"].max()].copy()
        merge_keys = ["reit", "macro_market", "beds"]
        available_sp_cols = [c for c in sp_cols if c in sp_latest.columns]
        sp_subset = sp_latest[merge_keys + available_sp_cols].copy()
        grp = grp.drop(columns=available_sp_cols, errors="ignore")
        grp = pd.merge(grp, sp_subset, on=merge_keys, how="left")

    # Ensure all columns present and ordered
    for c in SUMMARY_COLS:
        if c not in grp.columns:
            grp[c] = None

    return grp[SUMMARY_COLS]


def update_summary_history(existing_history_df, new_summary_df):
    """
    Append new_summary_df rows to existing_history_df, skipping dates already present.
    Returns the updated DataFrame.
    """
    if existing_history_df is None or existing_history_df.empty:
        print(f"  Summary history: creating new with {len(new_summary_df):,} rows")
        return new_summary_df.copy()

    existing_dates = set(existing_history_df["scrape_date"].astype(str).unique())
    new_date = str(new_summary_df["scrape_date"].iloc[0])

    if new_date in existing_dates:
        print(f"  Summary history: date {new_date} already exists, skipping append")
        return existing_history_df.copy()

    combined = pd.concat([existing_history_df, new_summary_df], ignore_index=True)
    print(f"  Summary history: appended {len(new_summary_df):,} rows for {new_date} "
          f"(total: {len(combined):,} rows)")
    return combined


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def set_col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def write_header_row(ws, row_num, headers, fill=None):
    fill = fill or HEADER_FILL
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.font   = HEADER_FONT
        cell.fill   = fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def write_data_row(ws, row_num, values, alt=False, number_formats=None):
    fill = PatternFill("solid", fgColor=ALT_ROW) if alt else None
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font      = BODY_FONT
        cell.alignment = BODY_ALIGN
        cell.border    = THIN_BORDER
        if fill:
            cell.fill = fill
        if number_formats and col_idx <= len(number_formats) and number_formats[col_idx - 1]:
            cell.number_format = number_formats[col_idx - 1]


def freeze_top_row(ws):
    ws.freeze_panes = "A2"


def add_title(ws, title, row=1):
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1: Inputs
# ─────────────────────────────────────────────────────────────────────────────

def build_inputs_sheet(wb, df, registry_df, latest_date_str):
    ws = wb.create_sheet("Inputs")
    ws.sheet_view.showGridLines = False

    add_title(ws, "REIT Rental Analysis — Configuration & Summary", row=1)

    ws.cell(row=3, column=1, value="Run Date").font = SUBHEAD_FONT
    ws.cell(row=3, column=2, value=pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"))

    ws.cell(row=4, column=1, value="Latest Scrape Date").font = SUBHEAD_FONT
    ws.cell(row=4, column=2, value=latest_date_str)

    ws.cell(row=5, column=1, value="Data Source").font = SUBHEAD_FONT
    ws.cell(row=5, column=2, value=f"github.com/{GITHUB_OWNER}/{GITHUB_REPO}/{DATA_PATH}")

    ws.cell(row=6, column=1, value="Total Rows").font = SUBHEAD_FONT
    ws.cell(row=6, column=2, value=len(df))

    ws.cell(row=7, column=1, value="REITs Covered").font = SUBHEAD_FONT
    ws.cell(row=7, column=2, value=", ".join(sorted(df["reit"].dropna().unique())))

    ws.cell(row=8, column=1, value="Scrape Dates").font = SUBHEAD_FONT
    dates_str = ", ".join(str(d)[:10] for d in sorted(df["scrape_date"].dropna().unique()))
    ws.cell(row=8, column=2, value=dates_str)

    ws.cell(row=10, column=1, value="REIT Coverage Summary").font = SUBHEAD_FONT
    headers = ["REIT", "Total Units", "Markets", "Communities",
               "Avg Rent ($)", "Concession Rate (%)"]
    write_header_row(ws, 11, headers)

    reit_grp = df.groupby("reit")
    for i, (reit, grp) in enumerate(reit_grp):
        alt = (i % 2 == 1)
        avg_rent = grp["rent"].mean()
        conc_rate = grp["has_concession"].mean() if "has_concession" in grp else float("nan")
        row_vals = [
            reit,
            len(grp),
            grp["market"].nunique() if "market" in grp else "",
            grp["community"].nunique() if "community" in grp else "",
            round(avg_rent, 0) if pd.notna(avg_rent) else "",
            round(conc_rate, 4) if pd.notna(conc_rate) else "",
        ]
        fmts = [None, NUM_COMMA, NUM_COMMA, NUM_COMMA, NUM_CURRENCY, NUM_PCT]
        write_data_row(ws, 12 + i, row_vals, alt=alt, number_formats=fmts)

    freeze_top_row(ws)
    set_col_widths(ws, {"A": 28, "B": 50, "C": 18, "D": 18, "E": 16, "F": 20})
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2: Data (latest week raw panel)
# ─────────────────────────────────────────────────────────────────────────────

def _write_raw_data_sheet(wb, df, sheet_name, header_fill=None):
    """Shared helper to write a raw data sheet (Data or Data_Prior)."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    fill = header_fill or HEADER_FILL

    display_cols = [c for c in EXPECTED_COLS + ["macro_market"] if c in df.columns]
    write_header_row(ws, 1, display_cols, fill=fill)

    fmts = []
    for c in display_cols:
        if c in ("rent", "effective_monthly_rent", "sqft"):
            fmts.append(NUM_CURRENCY if "rent" in c else NUM_COMMA)
        elif c in ("has_concession",):
            fmts.append(None)
        else:
            fmts.append(None)

    for i, row in enumerate(df[display_cols].itertuples(index=False), start=2):
        alt = (i % 2 == 0)
        vals = list(row)
        vals_clean = [
            v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else
            (None if (isinstance(v, float) and pd.isna(v)) else v)
            for v in vals
        ]
        write_data_row(ws, i, vals_clean, alt=alt, number_formats=fmts)

    freeze_top_row(ws)

    widths = {}
    for idx, col in enumerate(display_cols, start=1):
        ltr = get_column_letter(idx)
        widths[ltr] = max(12, len(col) + 2)
    set_col_widths(ws, widths)
    return ws


def build_data_sheet(wb, df):
    """Build Data sheet with latest week only."""
    latest = df["scrape_date"].max()
    df_latest = df[df["scrape_date"] == latest].copy()
    print(f"    Data sheet: {len(df_latest):,} rows (latest period: {str(latest)[:10]})")
    return _write_raw_data_sheet(wb, df_latest, "Data")


def build_data_prior_sheet(wb, df):
    """Build Data_Prior sheet with the prior week's raw data."""
    dates = sorted(df["scrape_date"].dropna().unique())
    if len(dates) < 2:
        ws = wb.create_sheet("Data_Prior")
        ws.sheet_view.showGridLines = False
        ws.cell(row=1, column=1,
                value="No prior period data yet — run again after Week 2 scrape").font = Font(
            name="Arial", italic=True, color="888888", size=10)
        print("    Data_Prior sheet: no prior period data")
        return ws
    prior_date = dates[-2]
    df_prior = df[df["scrape_date"] == prior_date].copy()
    # Gray header to distinguish from current
    prior_fill = PatternFill("solid", fgColor="607D8B")
    print(f"    Data_Prior sheet: {len(df_prior):,} rows (prior period: {str(prior_date)[:10]})")
    return _write_raw_data_sheet(wb, df_prior, "Data_Prior", header_fill=prior_fill)


# ─────────────────────────────────────────────────────────────────────────────
# SHEET: Summary_History (persistent time series)
# ─────────────────────────────────────────────────────────────────────────────

def build_summary_history_sheet(wb, summary_history_df):
    """Write full summary_history time series to a sheet."""
    ws = wb.create_sheet("Summary_History")
    ws.sheet_view.showGridLines = False
    add_title(ws, "Summary History — All Periods", row=1)

    if summary_history_df is None or summary_history_df.empty:
        ws.cell(row=3, column=1, value="No summary history data yet.").font = Font(
            name="Arial", italic=True, color="888888", size=10)
        return ws

    headers = list(summary_history_df.columns)
    write_header_row(ws, 2, headers)

    fmts = []
    for c in headers:
        if c in ("avg_rent", "median_rent", "avg_concession_value",
                 "sp_avg_rent_curr", "sp_avg_rent_prev",
                 "avg_eff_rent", "sp_avg_eff_rent_curr", "sp_avg_eff_rent_prev"):
            fmts.append(NUM_CURRENCY)
        elif c in ("avg_rent_psf", "median_rent_psf", "avg_eff_rent_psf",
                    "sp_avg_rent_psf_curr", "sp_avg_rent_psf_prev",
                    "sp_avg_eff_rent_psf_curr", "sp_avg_eff_rent_psf_prev",
                    "rent_per_sqft"):
            fmts.append('$#,##0.00')
        elif "pct" in c.lower() or "rate" in c.lower():
            fmts.append(NUM_PCT)
        elif c in ("listing_count", "avg_sqft", "sp_count"):
            fmts.append(NUM_COMMA)
        else:
            fmts.append(None)

    for i, row in enumerate(summary_history_df.itertuples(index=False), start=3):
        alt = (i % 2 == 0)
        vals = [
            v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else
            (None if (isinstance(v, float) and pd.isna(v)) else v)
            for v in row
        ]
        write_data_row(ws, i, vals, alt=alt, number_formats=fmts)

    freeze_top_row(ws)
    widths = {}
    for idx, col in enumerate(headers, start=1):
        ltr = get_column_letter(idx)
        widths[ltr] = max(12, len(str(col)) + 2)
    set_col_widths(ws, widths)

    n_dates = summary_history_df["scrape_date"].nunique() if "scrape_date" in summary_history_df.columns else 0
    print(f"    Summary_History sheet: {len(summary_history_df):,} rows, {n_dates} period(s)")
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3: Market_Calcs
# ─────────────────────────────────────────────────────────────────────────────

def build_market_calcs_sheet(wb, df):
    ws = wb.create_sheet("Market_Calcs")
    ws.sheet_view.showGridLines = False
    add_title(ws, "Market-Level Calculations", row=1)

    grp_cols = ["scrape_date", "macro_market", "beds"]
    agg = (
        df.groupby(grp_cols, dropna=False)
        .agg(
            unit_count=("unit_id", "count"),
            avg_rent=("rent", "mean"),
            median_rent=("rent", "median"),
            avg_sqft=("sqft", "mean"),
            concession_rate=("has_concession", "mean"),
            avg_eff_rent=("effective_monthly_rent", "mean"),
        )
        .reset_index()
        .sort_values(["scrape_date", "macro_market", "beds"])
    )

    headers = ["Scrape Date", "Macro Market", "Beds", "Unit Count",
               "Avg Rent", "Median Rent", "Avg Sqft",
               "Concession Rate", "Avg Eff Rent"]
    write_header_row(ws, 2, headers)
    fmts = [None, None, None, NUM_COMMA,
            NUM_CURRENCY, NUM_CURRENCY, NUM_COMMA, NUM_PCT, NUM_CURRENCY]

    for i, row in enumerate(agg.itertuples(index=False), start=3):
        alt = (i % 2 == 0)
        vals = [
            str(row.scrape_date)[:10] if pd.notna(row.scrape_date) else "",
            row.macro_market,
            int(row.beds) if pd.notna(row.beds) else "N/A",
            row.unit_count,
            round(row.avg_rent, 0) if pd.notna(row.avg_rent) else None,
            round(row.median_rent, 0) if pd.notna(row.median_rent) else None,
            round(row.avg_sqft, 0) if pd.notna(row.avg_sqft) else None,
            round(row.concession_rate, 4) if pd.notna(row.concession_rate) else None,
            round(row.avg_eff_rent, 0) if pd.notna(row.avg_eff_rent) else None,
        ]
        write_data_row(ws, i, vals, alt=alt, number_formats=fmts)

    freeze_top_row(ws)
    set_col_widths(ws, {"A": 14, "B": 26, "C": 8, "D": 12,
                        "E": 12, "F": 14, "G": 10, "H": 18, "I": 14})
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4: REIT_Summary
# ─────────────────────────────────────────────────────────────────────────────

def build_reit_summary_sheet(wb, df):
    ws = wb.create_sheet("REIT_Summary")
    ws.sheet_view.showGridLines = False
    add_title(ws, "REIT-Level Summary (Latest Scrape Date)", row=1)

    latest = df["scrape_date"].max()
    df_latest = df[df["scrape_date"] == latest].copy()

    grp = (
        df_latest.groupby(["reit", "beds"], dropna=False)
        .agg(
            unit_count=("unit_id", "count"),
            avg_rent=("rent", "mean"),
            median_rent=("rent", "median"),
            avg_sqft=("sqft", "mean"),
            concession_rate=("has_concession", "mean"),
            avg_eff_rent=("effective_monthly_rent", "mean"),
            markets=("macro_market", "nunique"),
        )
        .reset_index()
        .sort_values(["reit", "beds"])
    )

    headers = ["REIT", "Beds", "Unit Count", "Avg Rent", "Median Rent",
               "Avg Sqft", "Concession Rate", "Avg Eff Rent", "# Markets"]
    write_header_row(ws, 2, headers)
    fmts = [None, None, NUM_COMMA, NUM_CURRENCY, NUM_CURRENCY,
            NUM_COMMA, NUM_PCT, NUM_CURRENCY, None]

    for i, row in enumerate(grp.itertuples(index=False), start=3):
        alt = (i % 2 == 0)
        vals = [
            row.reit,
            int(row.beds) if pd.notna(row.beds) else "N/A",
            row.unit_count,
            round(row.avg_rent, 0) if pd.notna(row.avg_rent) else None,
            round(row.median_rent, 0) if pd.notna(row.median_rent) else None,
            round(row.avg_sqft, 0) if pd.notna(row.avg_sqft) else None,
            round(row.concession_rate, 4) if pd.notna(row.concession_rate) else None,
            round(row.avg_eff_rent, 0) if pd.notna(row.avg_eff_rent) else None,
            row.markets,
        ]
        write_data_row(ws, i, vals, alt=alt, number_formats=fmts)

    freeze_top_row(ws)
    set_col_widths(ws, {"A": 12, "B": 8, "C": 12, "D": 12, "E": 14,
                        "F": 10, "G": 18, "H": 14, "I": 12})
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5: Macro_Market_Detail
# ─────────────────────────────────────────────────────────────────────────────

def build_macro_market_detail_sheet(wb, df):
    ws = wb.create_sheet("Macro_Market_Detail")
    ws.sheet_view.showGridLines = False
    add_title(ws, "Macro Market Detail — REIT × Market × Beds", row=1)

    latest = df["scrape_date"].max()
    df_latest = df[df["scrape_date"] == latest].copy()

    grp = (
        df_latest.groupby(["macro_market", "reit", "beds"], dropna=False)
        .agg(
            unit_count=("unit_id", "count"),
            avg_rent=("rent", "mean"),
            median_rent=("rent", "median"),
            concession_rate=("has_concession", "mean"),
            avg_eff_rent=("effective_monthly_rent", "mean"),
        )
        .reset_index()
        .sort_values(["macro_market", "reit", "beds"])
    )

    headers = ["Macro Market", "REIT", "Beds", "Unit Count",
               "Avg Rent", "Median Rent", "Concession Rate", "Avg Eff Rent"]
    write_header_row(ws, 2, headers)
    fmts = [None, None, None, NUM_COMMA,
            NUM_CURRENCY, NUM_CURRENCY, NUM_PCT, NUM_CURRENCY]

    for i, row in enumerate(grp.itertuples(index=False), start=3):
        alt = (i % 2 == 0)
        vals = [
            row.macro_market,
            row.reit,
            int(row.beds) if pd.notna(row.beds) else "N/A",
            row.unit_count,
            round(row.avg_rent, 0) if pd.notna(row.avg_rent) else None,
            round(row.median_rent, 0) if pd.notna(row.median_rent) else None,
            round(row.concession_rate, 4) if pd.notna(row.concession_rate) else None,
            round(row.avg_eff_rent, 0) if pd.notna(row.avg_eff_rent) else None,
        ]
        write_data_row(ws, i, vals, alt=alt, number_formats=fmts)

    freeze_top_row(ws)
    set_col_widths(ws, {"A": 26, "B": 12, "C": 8, "D": 12,
                        "E": 12, "F": 14, "G": 18, "H": 14})
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 6: Charts_Rent
# ─────────────────────────────────────────────────────────────────────────────

def build_charts_rent_sheet(wb, df):
    ws = wb.create_sheet("Charts_Rent")
    ws.sheet_view.showGridLines = False
    add_title(ws, "Average Rent by REIT and Bed Type", row=1)

    # Build a pivot: rows = beds, cols = reit, values = avg_rent (latest date)
    latest = df["scrape_date"].max()
    df_latest = df[df["scrape_date"] == latest].copy()

    pivot = df_latest.groupby(["reit", "beds"])["rent"].mean().unstack("beds")
    pivot = pivot.round(0)

    reits = list(pivot.index)
    beds_vals = sorted([c for c in pivot.columns if pd.notna(c)])

    # Write table starting at row 3
    hdr = ["REIT"] + [f"{int(b)}BR" if b != 0 else "Studio" for b in beds_vals]
    write_header_row(ws, 3, hdr)

    for i, reit in enumerate(reits):
        vals = [reit] + [
            pivot.loc[reit, b] if b in pivot.columns and pd.notna(pivot.loc[reit, b]) else None
            for b in beds_vals
        ]
        fmts = [None] + [NUM_CURRENCY] * len(beds_vals)
        write_data_row(ws, 4 + i, vals, alt=(i % 2 == 1), number_formats=fmts)

    # Bar chart
    if reits and beds_vals:
        chart = BarChart()
        chart.type    = "col"
        chart.grouping = "clustered"
        chart.title   = f"Avg Rent by REIT & Bed Type ({str(latest)[:10]})"
        chart.y_axis.title = "Avg Monthly Rent ($)"
        chart.x_axis.title = "REIT"
        chart.style  = 10
        chart.width  = 24
        chart.height = 14

        data_ref = Reference(ws, min_col=2, max_col=1 + len(beds_vals),
                             min_row=3, max_row=3 + len(reits))
        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(reits))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        # Color each series by bed type (not REIT-based, use fallback)
        ws.add_chart(chart, f"A{5 + len(reits)}")

    freeze_top_row(ws)
    set_col_widths(ws, {get_column_letter(i): 14 for i in range(1, 2 + len(beds_vals))})
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 7: Charts_Concessions
# ─────────────────────────────────────────────────────────────────────────────

def build_charts_concessions_sheet(wb, df):
    ws = wb.create_sheet("Charts_Concessions")
    ws.sheet_view.showGridLines = False
    add_title(ws, "Concession Rate by REIT and Date", row=1)

    grp = (
        df.groupby(["scrape_date", "reit"])["has_concession"]
        .mean()
        .reset_index()
        .rename(columns={"has_concession": "concession_rate"})
    )
    grp["scrape_date"] = grp["scrape_date"].apply(lambda x: str(x)[:10])

    dates_sorted = sorted(grp["scrape_date"].unique())
    reits_sorted = sorted(grp["reit"].unique())

    pivot = grp.pivot(index="scrape_date", columns="reit", values="concession_rate")
    pivot = pivot.round(4)

    # Write table
    hdr = ["Date"] + reits_sorted
    write_header_row(ws, 3, hdr)

    for i, dt in enumerate(dates_sorted):
        row_vals = [dt] + [
            pivot.loc[dt, r] if r in pivot.columns and dt in pivot.index and pd.notna(pivot.loc[dt, r]) else None
            for r in reits_sorted
        ]
        fmts = [None] + [NUM_PCT] * len(reits_sorted)
        write_data_row(ws, 4 + i, row_vals, alt=(i % 2 == 1), number_formats=fmts)

    # Line chart
    if dates_sorted and reits_sorted:
        chart = LineChart()
        chart.title   = "Concession Rate Over Time by REIT"
        chart.y_axis.title = "Concession Rate"
        chart.x_axis.title = "Scrape Date"
        chart.style  = 10
        chart.width  = 24
        chart.height = 14

        data_ref = Reference(ws, min_col=2, max_col=1 + len(reits_sorted),
                             min_row=3, max_row=3 + len(dates_sorted))
        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(dates_sorted))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        _apply_series_colors(chart, reits_sorted, REIT_COLORS)
        ws.add_chart(chart, f"A{5 + len(dates_sorted)}")

    freeze_top_row(ws)
    set_col_widths(ws, {get_column_letter(i): 14 for i in range(1, 2 + len(reits_sorted))})
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 8: Same_Prop_Trends
# ─────────────────────────────────────────────────────────────────────────────

def _build_sp_index_from_history(summary_history_df, sp_df, metric_curr, metric_prev):
    """
    Build a pivot index (base=100) from summary_history or sp_df for a given metric.
    Returns (pivot_index_df, reits_list) or (None, []).
    """
    pivot_index = None
    reits = sorted(sp_df["reit"].dropna().unique()) if not sp_df.empty else []

    if (summary_history_df is not None and not summary_history_df.empty
            and metric_curr in summary_history_df.columns):
        hist = summary_history_df.copy()
        hist_sp = hist.dropna(subset=[metric_curr])
        if not hist_sp.empty:
            hist_rent = (
                hist_sp.groupby(["reit", "scrape_date"])[metric_curr]
                .mean().reset_index()
                .rename(columns={"scrape_date": "date", metric_curr: "val"})
            )
            hist_rent = hist_rent.sort_values("date")
            pivot_raw = hist_rent.pivot(index="date", columns="reit", values="val")
            # Only use summary_history if it has >= 2 dates; otherwise fall through to sp_df
            if len(pivot_raw) >= 2:
                pivot_index = pivot_raw.copy()
                for col in pivot_index.columns:
                    first_val = pivot_index[col].dropna().iloc[0] if not pivot_index[col].dropna().empty else None
                    if first_val and first_val != 0:
                        pivot_index[col] = (pivot_index[col] / first_val * 100).round(2)
                return pivot_index, list(pivot_index.columns)

    if not sp_df.empty and metric_curr in sp_df.columns and metric_prev in sp_df.columns:
        first_prev = sp_df["date_prev"].min()
        base = (
            sp_df[sp_df["date_prev"] == first_prev]
            .groupby("reit")[metric_prev].mean().reset_index()
            .rename(columns={metric_prev: "val"})
        )
        base["date"] = first_prev
        curr_part = (
            sp_df.groupby(["reit", "date_curr"])[metric_curr].mean().reset_index()
            .rename(columns={"date_curr": "date", metric_curr: "val"})
        )
        combined = pd.concat([base, curr_part], ignore_index=True).sort_values("date")
        pivot_raw = combined.pivot(index="date", columns="reit", values="val")
        pivot_index = pivot_raw.copy()
        for col in pivot_index.columns:
            first_val = pivot_index[col].dropna().iloc[0] if not pivot_index[col].dropna().empty else None
            if first_val and first_val != 0:
                pivot_index[col] = (pivot_index[col] / first_val * 100).round(2)
        return pivot_index, reits

    return None, reits


def _write_index_section(ws, start_row, title, calc_desc, pivot_index, chart_title, legend_font):
    """
    Write an index table + line chart to the worksheet. Returns the row after the chart.
    """
    from openpyxl.comments import Comment

    ws.cell(row=start_row - 1, column=1, value=title).font = SUBHEAD_FONT
    ws.cell(row=start_row - 1, column=4, value=calc_desc).font = legend_font

    if pivot_index is None or pivot_index.empty:
        ws.cell(row=start_row, column=1,
                value="No data available for this metric.").font = Font(
            name="Arial", italic=True, color="888888", size=10)
        return start_row + 2

    idx_hdr = ["Date"] + list(pivot_index.columns)
    write_header_row(ws, start_row, idx_hdr)
    for col_idx in range(2, len(idx_hdr) + 1):
        ws.cell(row=start_row, column=col_idx).comment = Comment(
            "Index = (this period value / base period value) x 100. "
            "100 = no change from base. Computed on same-property pool only.", "build_excel.py")

    for i, (dt, row_series) in enumerate(pivot_index.iterrows()):
        r = start_row + 1 + i
        row_vals = [str(dt)[:10]] + [
            float(row_series[c]) if pd.notna(row_series.get(c)) else None
            for c in pivot_index.columns
        ]
        write_data_row(ws, r, row_vals, alt=(i % 2 == 1))

    n_idx_rows = len(pivot_index)

    if n_idx_rows >= 2 and len(pivot_index.columns) > 0:
        line_chart = LineChart()
        line_chart.title = chart_title
        line_chart.y_axis.title = "Index"
        line_chart.x_axis.title = "Scrape Date"
        line_chart.style = 10
        line_chart.width = 24
        line_chart.height = 14

        data_ref = Reference(ws, min_col=2, max_col=1 + len(pivot_index.columns),
                             min_row=start_row, max_row=start_row + n_idx_rows)
        cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n_idx_rows)
        line_chart.add_data(data_ref, titles_from_data=True)
        line_chart.set_categories(cats_ref)
        _apply_series_colors(line_chart, list(pivot_index.columns), REIT_COLORS)
        ws.add_chart(line_chart, f"A{start_row + n_idx_rows + 2}")
        return start_row + n_idx_rows + 2 + 18  # chart height ~18 rows
    else:
        ws.cell(row=start_row + n_idx_rows + 2, column=1,
                value="Trends available from Week 2+ (need >= 2 data points for a line chart).").font = Font(
            name="Arial", italic=True, color="888888", size=10)
        return start_row + n_idx_rows + 4


def build_same_prop_sheet(wb, df, sp_df, summary_history_df=None):
    """
    Build Same_Prop_Trends with hardcoded values + methodology comments.
    Uses summary_history_df for the rent index chart (full history).
    """
    ws = wb.create_sheet("Same_Prop_Trends")
    ws.sheet_view.showGridLines = False

    dates = sorted(df["scrape_date"].dropna().unique())

    if len(dates) < 2 or sp_df.empty:
        add_title(ws, "Same-Property Trends", row=1)
        ws.cell(row=3, column=1,
                value="Insufficient data — run again after Week 2 scrape").font = Font(
            name="Arial", italic=True, color="888888", size=10)
        return ws

    date_strs = [d.strftime("%Y-%m-%d") for d in dates]

    add_title(ws, "Same-Property WoW Rent Change — REIT x Macro Market x Beds", row=1)

    # ── Methodology legend ──
    legend_font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.merge_cells("A2:V2")
    ws.cell(row=2, column=1,
            value="METHODOLOGY: Same-property pool = unit_ids present in BOTH the current and prior scrape period. "
                  "All metrics below are computed only on this matched pool to isolate true rent changes from mix shifts.").font = legend_font

    # Column methodology comments (row 3 headers get comments)
    from openpyxl.comments import Comment
    col_comments = {
        4: "CALC: Count of unit_ids present in both current and prior period, for this REIT / macro-market / bed-count.",
        5: "CALC: Mean asking rent of the same-property pool in the CURRENT period (latest scrape date).",
        6: "CALC: Mean asking rent of the same-property pool in the PRIOR period (previous scrape date). Same exact units as col E.",
        7: "CALC: (Avg Rent Curr - Avg Rent Prev) / Avg Rent Prev. Green >+0.5%, Red <-0.5%.",
        8: "CALC: Share of same-property units offering a concession in the CURRENT period. = count(has_concession=True) / total count.",
        9: "CALC: Share of same-property units offering a concession in the PRIOR period.",
        10: "CALC: Mean rent/sqft for same-property units (sqft>0) in CURRENT period.",
        11: "CALC: Mean rent/sqft for same-property units (sqft>0) in PRIOR period.",
        12: "CALC: (Rent PSF Curr - Rent PSF Prev) / Rent PSF Prev. Green >+0.5%, Red <-0.5%.",
        13: "CALC: Mean effective_monthly_rent for same-property units (non-null) in CURRENT period.",
        14: "CALC: Mean effective_monthly_rent for same-property units (non-null) in PRIOR period.",
        15: "CALC: (Eff Rent Curr - Eff Rent Prev) / Eff Rent Prev.",
        16: "CALC: Mean effective_monthly_rent/sqft for same-property units in CURRENT period.",
        17: "CALC: Mean effective_monthly_rent/sqft for same-property units in PRIOR period.",
        18: "CALC: (Eff PSF Curr - Eff PSF Prev) / Eff PSF Prev.",
    }

    headers = ["REIT", "Macro Market", "Beds", "Same-Prop Count",
               "Avg Rent (Curr)", "Avg Rent (Prev)", "WoW Chg (%)",
               "Concession Rate (Curr)", "Concession Rate (Prev)",
               "Rent PSF (Curr)", "Rent PSF (Prev)", "WoW PSF (%)",
               "Eff Rent (Curr)", "Eff Rent (Prev)", "WoW Eff (%)",
               "Eff PSF (Curr)", "Eff PSF (Prev)", "WoW Eff PSF (%)",
               "Period"]
    write_header_row(ws, 3, headers)

    for col_idx, comment_text in col_comments.items():
        ws.cell(row=3, column=col_idx).comment = Comment(comment_text, "build_excel.py")

    green_fill = PatternFill("solid", fgColor=LT_GREEN)
    red_fill   = PatternFill("solid", fgColor=LT_RED)

    fmts = [None, None, None, NUM_COMMA,
            NUM_CURRENCY, NUM_CURRENCY, NUM_PCT,
            NUM_PCT, NUM_PCT,
            '$#,##0.00', '$#,##0.00', NUM_PCT,
            NUM_CURRENCY, NUM_CURRENCY, NUM_PCT,
            '$#,##0.00', '$#,##0.00', NUM_PCT,
            None]

    latest_curr = sp_df["date_curr"].max()
    sp_latest = sp_df[sp_df["date_curr"] == latest_curr].copy()

    data_rows = sorted(
        sp_latest.itertuples(index=False),
        key=lambda r: (str(r.reit), str(r.macro_market), r.beds if pd.notna(r.beds) else 99)
    )

    wow_pct_cols = [7, 12, 15, 18]  # columns with WoW %

    for i, row in enumerate(data_rows):
        excel_row = 4 + i
        alt = (i % 2 == 1)
        beds_disp = int(row.beds) if pd.notna(row.beds) else "N/A"
        period_str = f"{str(row.date_prev)[:10]} -> {str(row.date_curr)[:10]}"

        def _rv(attr, ndigits=0):
            v = getattr(row, attr, None)
            if v is not None and pd.notna(v):
                return round(float(v), ndigits)
            return None

        vals = [
            row.reit,
            row.macro_market,
            beds_disp,
            int(row.sp_count) if pd.notna(row.sp_count) else None,
            _rv("sp_avg_rent_curr", 0),
            _rv("sp_avg_rent_prev", 0),
            _rv("sp_wow_pct", 4),
            _rv("sp_concession_rate_curr", 4),
            _rv("sp_concession_rate_prev", 4),
            _rv("sp_avg_rent_psf_curr", 2),
            _rv("sp_avg_rent_psf_prev", 2),
            _rv("sp_wow_pct_psf", 4),
            _rv("sp_avg_eff_rent_curr", 0),
            _rv("sp_avg_eff_rent_prev", 0),
            _rv("sp_wow_pct_eff", 4),
            _rv("sp_avg_eff_rent_psf_curr", 2),
            _rv("sp_avg_eff_rent_psf_prev", 2),
            _rv("sp_wow_pct_eff_psf", 4),
            period_str,
        ]
        write_data_row(ws, excel_row, vals, alt=alt, number_formats=fmts)

        # Highlight WoW columns
        for wow_col in wow_pct_cols:
            attr_map = {7: "sp_wow_pct", 12: "sp_wow_pct_psf", 15: "sp_wow_pct_eff", 18: "sp_wow_pct_eff_psf"}
            wow_val = getattr(row, attr_map[wow_col], None)
            if wow_val is not None and pd.notna(wow_val):
                cell = ws.cell(row=excel_row, column=wow_col)
                if wow_val > 0.005:
                    cell.fill = green_fill
                elif wow_val < -0.005:
                    cell.fill = red_fill

    n_data = len(data_rows)

    # ── Same-Property Avg Rent by REIT (raw $) — from sp_df (current 2 weeks) ──
    rent_start = 4 + n_data + 3
    ws.cell(row=rent_start - 1, column=1,
            value="Same-Property Avg Rent by REIT ($)").font = SUBHEAD_FONT
    ws.cell(row=rent_start - 1, column=4,
            value="CALC: For each REIT and date, avg asking rent across all same-property units "
                  "(unit_ids present in both periods), all markets and bed counts.").font = legend_font

    reits = sorted(sp_df["reit"].dropna().unique())

    # Build avg rent per REIT per date from sp_df (current 2-week window)
    first_prev = sp_df["date_prev"].min()
    reit_date_rent_base = (
        sp_df[sp_df["date_prev"] == first_prev]
        .groupby("reit")["sp_avg_rent_prev"].mean().reset_index()
        .rename(columns={"sp_avg_rent_prev": "avg_rent"})
    )
    reit_date_rent_base["date"] = first_prev

    reit_date_rent_curr = (
        sp_df.groupby(["reit", "date_curr"])["sp_avg_rent_curr"].mean().reset_index()
        .rename(columns={"date_curr": "date", "sp_avg_rent_curr": "avg_rent"})
    )

    combined_rent = pd.concat([reit_date_rent_base, reit_date_rent_curr], ignore_index=True)
    combined_rent = combined_rent.sort_values("date")
    pivot_rent = combined_rent.pivot(index="date", columns="reit", values="avg_rent")

    rent_hdr = ["Date"] + list(pivot_rent.columns)
    write_header_row(ws, rent_start, rent_hdr)
    for col_idx in range(2, len(rent_hdr) + 1):
        ws.cell(row=rent_start, column=col_idx).comment = Comment(
            "Avg asking rent ($) for this REIT's same-property pool on this date.", "build_excel.py")

    for i, (dt, row_series) in enumerate(pivot_rent.iterrows()):
        r = rent_start + 1 + i
        row_vals = [str(dt)[:10]] + [
            round(float(row_series[reit]), 0) if pd.notna(row_series.get(reit)) else None
            for reit in pivot_rent.columns
        ]
        rent_fmts = [None] + [NUM_CURRENCY] * len(reits)
        write_data_row(ws, r, row_vals, alt=(i % 2 == 1), number_formats=rent_fmts)

    n_rent_rows = len(pivot_rent)

    # ──────────────────────────────────────────────────────────────────────
    # INDEX SECTIONS — 4 total: Gross Rent, Gross Rent PSF, Net Eff Rent, Net Eff Rent PSF
    # ─────────────────────────────────���────────────────────────────────────

    # 1. Gross Asking Rent Index
    idx_start = rent_start + n_rent_rows + 3
    pi_rent, _ = _build_sp_index_from_history(
        summary_history_df, sp_df, "sp_avg_rent_curr", "sp_avg_rent_prev")
    next_row = _write_index_section(
        ws, idx_start,
        "Same-Property Avg Rent Index (Base = 100)",
        "CALC: (Avg rent on date N / Avg rent on base date) x 100. Uses FULL history from summary_history.csv.",
        pi_rent,
        "Same-Property Avg Rent Index (Base = 100)",
        legend_font,
    )

    # 2. Gross Asking Rent PSF Index
    pi_psf, _ = _build_sp_index_from_history(
        summary_history_df, sp_df, "sp_avg_rent_psf_curr", "sp_avg_rent_psf_prev")
    next_row = _write_index_section(
        ws, next_row + 2,
        "Same-Property Rent PSF Index (Base = 100)",
        "CALC: (Avg rent PSF on date N / Avg rent PSF on base date) x 100.",
        pi_psf,
        "Same-Property Rent PSF Index (Base = 100)",
        legend_font,
    )

    # 3. Net Effective Rent Index
    pi_eff, _ = _build_sp_index_from_history(
        summary_history_df, sp_df, "sp_avg_eff_rent_curr", "sp_avg_eff_rent_prev")
    next_row = _write_index_section(
        ws, next_row + 2,
        "Same-Property Net Effective Rent Index (Base = 100)",
        "CALC: (Avg effective rent on date N / Avg effective rent on base date) x 100.",
        pi_eff,
        "Same-Property Net Effective Rent Index (Base = 100)",
        legend_font,
    )

    # 4. Net Effective Rent PSF Index
    pi_eff_psf, _ = _build_sp_index_from_history(
        summary_history_df, sp_df, "sp_avg_eff_rent_psf_curr", "sp_avg_eff_rent_psf_prev")
    next_row = _write_index_section(
        ws, next_row + 2,
        "Same-Property Net Effective Rent PSF Index (Base = 100)",
        "CALC: (Avg effective rent PSF on date N / Avg effective rent PSF on base date) x 100.",
        pi_eff_psf,
        "Same-Property Net Effective Rent PSF Index (Base = 100)",
        legend_font,
    )

    freeze_top_row(ws)
    set_col_widths(ws, {
        "A": 12, "B": 26, "C": 8, "D": 16,
        "E": 16, "F": 16, "G": 12, "H": 22, "I": 22,
        "J": 14, "K": 14, "L": 12, "M": 14, "N": 14, "O": 12,
        "P": 14, "Q": 14, "R": 14, "S": 28,
    })
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# REGION ASSIGNMENTS (for per-REIT market breakdown charts)
# ─────────────────────────────────────────────────────────────────────────────

REGION_MAP = {}

_SOUTHEAST = ["Atlanta", "Charlotte", "Raleigh/Durham", "Nashville", "Charleston",
              "Jacksonville", "Savannah", "Greenville", "Birmingham", "Huntsville",
              "Memphis", "Chattanooga", "Richmond", "Hampton Roads", "Charlottesville", "Louisville"]
_SOUTHWEST = ["Dallas/Fort Worth", "Houston", "Austin", "San Antonio", "Phoenix",
              "Denver", "Las Vegas", "Salt Lake City", "Colorado Springs"]
_FLORIDA = ["Miami/Fort Lauderdale", "Orlando", "Tampa", "Gulf Shores", "Tallahassee",
            "Panama City Beach", "Gainesville"]
_EAST_COAST = ["New York", "Boston", "Washington, DC", "Philadelphia", "Baltimore",
               "Chicago", "Minneapolis", "Kansas City"]
_WEST_COAST = ["San Francisco", "San Francisco-East Bay", "San Jose", "Los Angeles",
               "Orange County", "San Diego", "Inland Empire", "Seattle", "Portland",
               "Santa Barbara", "Salinas/Monterey"]

for _mkt in _SOUTHEAST:
    REGION_MAP[_mkt] = "Southeast"
for _mkt in _SOUTHWEST:
    REGION_MAP[_mkt] = "Southwest"
for _mkt in _FLORIDA:
    REGION_MAP[_mkt] = "Florida"
for _mkt in _EAST_COAST:
    REGION_MAP[_mkt] = "East Coast"
for _mkt in _WEST_COAST:
    REGION_MAP[_mkt] = "West Coast"

REIT_LIST = ["MAA", "CPT", "EQR", "AVB", "UDR", "ESS", "INVH", "AMH"]


# ─────────────────────────────────────────────────────────────────────────────
# NEW SHEET: Per-REIT Market Breakdown
# ─────────────────────────────────────────────────────────────────────────────

def _compute_market_rent_psf_index(summary_history_df, sp_df, reit, market):
    """
    Compute indexed rent PSF (base=100) time series for a single REIT+market.
    Returns list of (date_str, index_value) or empty list.
    """
    if (summary_history_df is not None and not summary_history_df.empty
            and "sp_avg_rent_psf_curr" in summary_history_df.columns):
        hist = summary_history_df.copy()
        mask = (hist["reit"] == reit) & (hist["macro_market"] == market)
        sub = hist.loc[mask].dropna(subset=["sp_avg_rent_psf_curr"])
        if not sub.empty:
            by_date = sub.groupby("scrape_date")["sp_avg_rent_psf_curr"].mean().sort_index()
            if len(by_date) >= 2:
                base = by_date.iloc[0]
                if base and base != 0:
                    return [(str(d), round(v / base * 100, 2)) for d, v in by_date.items()]
    # Fallback to sp_df
    if not sp_df.empty and "sp_avg_rent_psf_curr" in sp_df.columns:
        sub = sp_df[(sp_df["reit"] == reit) & (sp_df["macro_market"] == market)]
        if not sub.empty:
            first_prev = sub["date_prev"].min()
            base_val = sub[sub["date_prev"] == first_prev]["sp_avg_rent_psf_prev"].mean()
            if pd.notna(base_val) and base_val != 0:
                result = [(str(first_prev)[:10], 100.0)]
                for _, grp in sub.groupby("date_curr"):
                    d = str(grp["date_curr"].iloc[0])[:10]
                    v = grp["sp_avg_rent_psf_curr"].mean()
                    if pd.notna(v):
                        result.append((d, round(v / base_val * 100, 2)))
                return result
    return []


def build_reit_market_sheets(wb, df, sp_df, summary_history_df):
    """
    For each REIT with data, create a {REIT}_Markets sheet with indexed rent PSF
    by market, grouped by region with line charts.
    """
    latest = df["scrape_date"].max()
    df_latest = df[df["scrape_date"] == latest].copy()
    legend_font = Font(name="Arial", italic=True, size=9, color="666666")

    dates = sorted(df["scrape_date"].dropna().unique())
    has_trends = len(dates) >= 2 and not sp_df.empty

    reits_with_data = sorted(df_latest["reit"].dropna().unique())
    target_reits = [r for r in REIT_LIST if r in reits_with_data]

    sheets_created = []

    for reit in target_reits:
        sheet_name = f"{reit}_Markets"
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        reit_data = df_latest[df_latest["reit"] == reit]
        # Markets for this REIT
        market_counts = reit_data.groupby("macro_market")["unit_id"].count().sort_values(ascending=False)
        reit_markets = list(market_counts.index)

        if not reit_markets:
            ws.cell(row=1, column=1, value=f"{reit} -- No market data available").font = TITLE_FONT
            sheets_created.append(sheet_name)
            continue

        add_title(ws, f"{reit} -- Same-Property Rent Index by Market (Base = 100)", row=1)

        if not has_trends:
            # Week 1 fallback: show current avg rent PSF by market
            ws.cell(row=3, column=1,
                    value="Trends available from Week 2+. Showing current avg rent PSF by market.").font = legend_font

            # Compute avg rent PSF per market
            reit_data_psf = reit_data.copy()
            reit_data_psf["_rpsf"] = reit_data_psf.apply(
                lambda r: r["rent"] / r["sqft"] if pd.notna(r["sqft"]) and r["sqft"] > 0 and pd.notna(r["rent"]) else None,
                axis=1,
            )
            mkt_psf = (
                reit_data_psf[reit_data_psf["_rpsf"].notna()]
                .groupby("macro_market")
                .agg(avg_rent_psf=("_rpsf", "mean"), count=("unit_id", "count"))
                .sort_values("count", ascending=False)
                .reset_index()
            )

            headers = ["Market", "Avg Rent PSF", "Listings"]
            write_header_row(ws, 4, headers)
            for i, row in enumerate(mkt_psf.itertuples(index=False)):
                write_data_row(ws, 5 + i,
                               [row.macro_market, round(row.avg_rent_psf, 2), row.count],
                               alt=(i % 2 == 1),
                               number_formats=[None, '$#,##0.00', NUM_COMMA])

            sheets_created.append(sheet_name)
            continue

        # Build index data for each market
        all_dates = set()
        market_index_data = {}  # market -> {date: index}
        for mkt in reit_markets:
            ts = _compute_market_rent_psf_index(summary_history_df, sp_df, reit, mkt)
            if ts:
                market_index_data[mkt] = {d: v for d, v in ts}
                all_dates.update(d for d, _ in ts)

        if not all_dates:
            ws.cell(row=3, column=1,
                    value="No same-property rent PSF data available yet.").font = legend_font
            sheets_created.append(sheet_name)
            continue

        sorted_dates = sorted(all_dates)
        markets_with_data = [m for m in reit_markets if m in market_index_data]

        # Write full table
        tbl_hdr = ["Date"] + markets_with_data
        write_header_row(ws, 3, tbl_hdr)
        for i, d in enumerate(sorted_dates):
            row_vals = [d] + [market_index_data[m].get(d) for m in markets_with_data]
            write_data_row(ws, 4 + i, row_vals, alt=(i % 2 == 1))

        n_tbl_rows = len(sorted_dates)

        # Charts by region
        current_row = 4 + n_tbl_rows + 2
        region_markets = {}
        for mkt in markets_with_data:
            region = REGION_MAP.get(mkt, "Other")
            region_markets.setdefault(region, []).append(mkt)

        for region_name in ["Southeast", "Southwest", "Florida", "East Coast", "West Coast", "Other"]:
            mkts = region_markets.get(region_name, [])
            if not mkts:
                continue

            # Top 5 by listing count
            mkt_order = [m for m in reit_markets if m in mkts and m in market_index_data][:5]
            if not mkt_order:
                continue

            # Write mini table for chart
            ws.cell(row=current_row, column=1,
                    value=f"{reit} -- {region_name} Markets").font = SUBHEAD_FONT
            current_row += 1

            mini_hdr = ["Date"] + mkt_order
            write_header_row(ws, current_row, mini_hdr)
            for i, d in enumerate(sorted_dates):
                row_vals = [d] + [market_index_data[m].get(d) for m in mkt_order]
                write_data_row(ws, current_row + 1 + i, row_vals, alt=(i % 2 == 1))

            n_mini = len(sorted_dates)

            if n_mini >= 2 and len(mkt_order) > 0:
                chart = LineChart()
                chart.title = f"{reit} -- {region_name} Markets, Same-Property Rent PSF Index"
                chart.y_axis.title = "Index"
                chart.x_axis.title = "Date"
                chart.style = 10
                chart.width = 24
                chart.height = 14

                data_ref = Reference(ws, min_col=2, max_col=1 + len(mkt_order),
                                     min_row=current_row, max_row=current_row + n_mini)
                cats_ref = Reference(ws, min_col=1,
                                     min_row=current_row + 1, max_row=current_row + n_mini)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                _apply_series_colors(chart, mkt_order, MARKET_COLORS)
                ws.add_chart(chart, f"A{current_row + n_mini + 2}")
                current_row = current_row + n_mini + 2 + 18
            else:
                current_row = current_row + n_mini + 3

        set_col_widths(ws, {get_column_letter(i): 16 for i in range(1, min(len(markets_with_data) + 2, 27))})
        sheets_created.append(sheet_name)

    return sheets_created


# ─────────────────────────────────────────────────────────────────────────────
# NEW SHEET: Market_Comparison
# ─────────────────────────────────────────────────────────────────────────────

def build_market_comparison_sheet(wb, df, sp_df, summary_history_df):
    """
    For each macro-market with >=2 REITs, show line chart comparing REITs
    on indexed rent PSF basis. Limited to top 15 markets by listing count.
    """
    ws = wb.create_sheet("Market_Comparison")
    ws.sheet_view.showGridLines = False
    legend_font = Font(name="Arial", italic=True, size=9, color="666666")

    add_title(ws, "Market Comparison -- Same-Property Rent PSF Index by Market", row=1)

    latest = df["scrape_date"].max()
    df_latest = df[df["scrape_date"] == latest].copy()

    dates = sorted(df["scrape_date"].dropna().unique())
    has_trends = len(dates) >= 2 and not sp_df.empty

    # Find markets with >= 2 REITs
    mkt_reit = df_latest.groupby("macro_market")["reit"].nunique()
    multi_reit_markets = mkt_reit[mkt_reit >= 2].index.tolist()

    # Sort by total listing count, take top 15
    mkt_counts = df_latest[df_latest["macro_market"].isin(multi_reit_markets)].groupby("macro_market")["unit_id"].count()
    mkt_counts = mkt_counts.sort_values(ascending=False)
    top_markets = list(mkt_counts.index[:15])

    if not top_markets:
        ws.cell(row=3, column=1,
                value="No markets with >= 2 REITs found.").font = Font(
            name="Arial", italic=True, color="888888", size=10)
        return ws

    current_row = 3

    if not has_trends:
        # Week 1 fallback: bar chart of current avg rent PSF by REIT per market
        ws.cell(row=current_row, column=1,
                value="Trends available from Week 2+. Showing current avg rent PSF by REIT per market.").font = legend_font
        current_row += 2

        for mkt in top_markets:
            mkt_data = df_latest[df_latest["macro_market"] == mkt].copy()
            mkt_data["_rpsf"] = mkt_data.apply(
                lambda r: r["rent"] / r["sqft"] if pd.notna(r["sqft"]) and r["sqft"] > 0 and pd.notna(r["rent"]) else None,
                axis=1,
            )
            reit_psf = (
                mkt_data[mkt_data["_rpsf"].notna()]
                .groupby("reit")["_rpsf"].mean()
                .sort_values(ascending=False)
                .reset_index()
            )
            if reit_psf.empty:
                continue

            ws.cell(row=current_row, column=1, value=mkt).font = SUBHEAD_FONT
            current_row += 1
            write_header_row(ws, current_row, ["REIT", "Avg Rent PSF"])
            for i, row in enumerate(reit_psf.itertuples(index=False)):
                write_data_row(ws, current_row + 1 + i,
                               [row.reit, round(row._rpsf, 2)],
                               alt=(i % 2 == 1),
                               number_formats=[None, '$#,##0.00'])

            n_reits = len(reit_psf)
            if n_reits > 0:
                chart = BarChart()
                chart.type = "col"
                chart.grouping = "clustered"
                chart.title = f"{mkt} -- Avg Rent PSF by REIT"
                chart.y_axis.title = "Rent PSF ($)"
                chart.style = 10
                chart.width = 24
                chart.height = 14

                data_ref = Reference(ws, min_col=2, max_col=2,
                                     min_row=current_row, max_row=current_row + n_reits)
                cats_ref = Reference(ws, min_col=1,
                                     min_row=current_row + 1, max_row=current_row + n_reits)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                # Bar chart with single series — color individual bars by REIT
                ws.add_chart(chart, f"D{current_row}")
                current_row += n_reits + 20

        set_col_widths(ws, {"A": 16, "B": 16})
        return ws

    # Normal case: line charts with indexed rent PSF
    for mkt in top_markets:
        # Find REITs in this market
        mkt_reits = sorted(df_latest[df_latest["macro_market"] == mkt]["reit"].unique())
        if len(mkt_reits) < 2:
            continue

        # Compute index for each REIT in this market
        all_dates = set()
        reit_index = {}  # reit -> {date: index}
        for reit in mkt_reits:
            ts = _compute_market_rent_psf_index(summary_history_df, sp_df, reit, mkt)
            if ts:
                reit_index[reit] = {d: v for d, v in ts}
                all_dates.update(d for d, _ in ts)

        reits_with_data = [r for r in mkt_reits if r in reit_index]
        if len(reits_with_data) < 2 or not all_dates:
            continue

        sorted_dates = sorted(all_dates)

        ws.cell(row=current_row, column=1, value=mkt).font = SUBHEAD_FONT
        current_row += 1

        tbl_hdr = ["Date"] + reits_with_data
        write_header_row(ws, current_row, tbl_hdr)
        for i, d in enumerate(sorted_dates):
            row_vals = [d] + [reit_index[r].get(d) for r in reits_with_data]
            write_data_row(ws, current_row + 1 + i, row_vals, alt=(i % 2 == 1))

        n_rows = len(sorted_dates)

        if n_rows >= 2 and len(reits_with_data) > 0:
            chart = LineChart()
            chart.title = f"{mkt} -- Same-Property Rent PSF Index by REIT"
            chart.y_axis.title = "Index"
            chart.x_axis.title = "Date"
            chart.style = 10
            chart.width = 24
            chart.height = 14

            data_ref = Reference(ws, min_col=2, max_col=1 + len(reits_with_data),
                                 min_row=current_row, max_row=current_row + n_rows)
            cats_ref = Reference(ws, min_col=1,
                                 min_row=current_row + 1, max_row=current_row + n_rows)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            _apply_series_colors(chart, reits_with_data, REIT_COLORS)
            ws.add_chart(chart, f"A{current_row + n_rows + 2}")
            current_row = current_row + n_rows + 2 + 18
        else:
            current_row = current_row + n_rows + 3

    set_col_widths(ws, {get_column_letter(i): 14 for i in range(1, 12)})
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  REIT Rental Analysis — Excel Workbook Builder")
    print("  Architecture: 2-week download + persistent summary_history")
    print("=" * 60)

    if not GITHUB_TOKEN:
        print("\n[WARNING] GITHUB_TOKEN is empty. Requests will be unauthenticated.")
        print("  Set GITHUB_TOKEN at the top of this script for private repos.\n")

    # Step 1: Download latest 2 weeks of CSVs from GitHub
    raw_files = fetch_latest_2_weeks_csvs()

    if not raw_files:
        print("[ERROR] No files downloaded. Check token, owner, repo, and DATA_PATH.")
        sys.exit(1)

    # Step 1b: Download existing summary_history.csv from GitHub
    existing_history = fetch_summary_history()

    # Step 1c: Download registry
    print("\n[Step 1b] Fetching unit registry...")
    registry_df = fetch_registry()

    # Step 2b: Merge parts
    print("\n[Step 2b] Merging part files...")
    merged_files = merge_parts(raw_files)

    # Step 3: Build panel (latest 2 weeks only)
    print("\n[Step 3] Building panel dataset...")
    df = build_panel(merged_files)

    # Step 4: Macro-market mapping
    print("\n[Step 4] Applying macro-market mapping...")
    df = apply_macro_map(df)

    # Step 5: Same-property analysis (2-week window)
    print("\n[Step 5] Computing same-property metrics...")
    sp_df = compute_same_property(df)

    # Step 5b: Build & update summary_history
    print("\n[Step 5b] Updating summary history...")
    current_summary = build_current_summary(df, sp_df)
    summary_history_df = update_summary_history(existing_history, current_summary)

    # Save summary_history locally
    save_summary_history(summary_history_df)

    # Step 6: Build Excel
    print("\n[Step 6] Building Excel workbook...")
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    latest_date = df["scrape_date"].max()
    latest_date_str = str(latest_date)[:10]

    print("  Building: Inputs")
    build_inputs_sheet(wb, df, registry_df, latest_date_str)

    print("  Building: Data (latest week only)")
    build_data_sheet(wb, df)

    print("  Building: Data_Prior (prior week)")
    build_data_prior_sheet(wb, df)

    print("  Building: Summary_History (all periods)")
    build_summary_history_sheet(wb, summary_history_df)

    print("  Building: Market_Calcs")
    build_market_calcs_sheet(wb, df)
    print("  Building: REIT_Summary")
    build_reit_summary_sheet(wb, df)
    print("  Building: Macro_Market_Detail")
    build_macro_market_detail_sheet(wb, df)
    print("  Building: Charts_Rent")
    build_charts_rent_sheet(wb, df)
    print("  Building: Charts_Concessions")
    build_charts_concessions_sheet(wb, df)
    print("  Building: Same_Prop_Trends")
    build_same_prop_sheet(wb, df, sp_df, summary_history_df=summary_history_df)

    print("  Building: Per-REIT Market sheets")
    reit_sheets = build_reit_market_sheets(wb, df, sp_df, summary_history_df)
    for s in reit_sheets:
        print(f"    Created: {s}")

    print("  Building: Market_Comparison")
    build_market_comparison_sheet(wb, df, sp_df, summary_history_df)

    # Step 7: Save Excel
    out_filename = f"REIT_Rental_Analysis_{latest_date_str}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_filename)
    out_path = os.path.normpath(out_path)
    wb.save(out_path)

    file_size_kb = os.path.getsize(out_path) / 1024

    # Print row counts for verification
    print(f"\n  --- Sheet Row Counts ---")
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"    {sname}: {ws.max_row:,} rows")

    print(f"\n{'=' * 60}")
    print(f"  Output saved: {out_path}")
    print(f"  File size:    {file_size_kb:.1f} KB")
    print(f"  Sheets:       {', '.join(wb.sheetnames)}")
    print(f"  Summary history: {LOCAL_SUMMARY_DIR}/summary_history.csv")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    main()
