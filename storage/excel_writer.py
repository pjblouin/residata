"""
Excel writer — appends weekly scrape results to a cumulative workbook.

Workbook layout (one sheet per purpose):
  raw_listings   — every unit row ever scraped (append-only)
  unit_registry  — current snapshot of unit_registry (overwrite each run)
  price_history  — one row per (unit_id, scrape_date, rent) change event
  concession_log — one row per (unit_id, scrape_date) where has_concession=True

On each run:
  1. Load existing workbook (or create fresh).
  2. Append new raw_listings rows (deduplicate by unit_id + scrape_date).
  3. Overwrite unit_registry sheet with latest registry CSV.
  4. Append to price_history only rows where rent changed vs. last seen.
  5. Append to concession_log only rows with has_concession=True.
  6. Save workbook.
"""

import logging
import os
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Formatting helpers ─────────────────────────────────────────────────────────

def _style_header_row(ws, row_num: int = 1) -> None:
    """Apply blue header styling to the given row."""
    for cell in ws[row_num]:
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


def _set_col_widths(ws, widths: dict[str, int]) -> None:
    """widths: {column_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Sheet writers ──────────────────────────────────────────────────────────────

def _upsert_sheet(wb: Workbook, name: str) -> object:
    """Return existing sheet or create it."""
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def _df_to_sheet_overwrite(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """Completely replace a sheet's content with df."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    _style_header_row(ws)
    _freeze_header(ws)


def _append_new_rows(
    wb: Workbook,
    sheet_name: str,
    new_df: pd.DataFrame,
    dedup_keys: list[str],
) -> int:
    """
    Append rows from new_df that aren't already in the sheet (by dedup_keys).
    Returns count of rows appended.
    """
    ws = _upsert_sheet(wb, sheet_name)

    # If sheet is empty (just created), write header + all rows
    if ws.max_row <= 1 and ws.cell(1, 1).value is None:
        for r in dataframe_to_rows(new_df, index=False, header=True):
            ws.append(r)
        _style_header_row(ws)
        _freeze_header(ws)
        return len(new_df)

    # Read existing keys from sheet
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    key_cols = {k: headers.index(k) for k in dedup_keys if k in headers}
    existing_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = tuple(row[key_cols[k]] for k in dedup_keys if k in key_cols)
        existing_keys.add(key)

    # Filter new_df to only genuinely new rows
    def row_key(r):
        return tuple(str(r[k]) if r[k] is not None else "" for k in dedup_keys)

    new_rows = new_df[~new_df.apply(row_key, axis=1).isin(
        {tuple(str(v) if v is not None else "" for v in k) for k in existing_keys}
    )]

    for r in dataframe_to_rows(new_rows, index=False, header=False):
        ws.append(r)

    return len(new_rows)


# ── Price history logic ────────────────────────────────────────────────────────

def _build_price_history_rows(new_df: pd.DataFrame, existing_ph: pd.DataFrame) -> pd.DataFrame:
    """
    Return only rows where rent differs from the unit's most recently recorded rent.
    On first appearance of a unit, always record it.
    """
    cols = ["unit_id", "scrape_date", "reit", "community", "market",
            "beds", "sqft", "rent", "effective_monthly_rent"]
    snapshot = new_df[[c for c in cols if c in new_df.columns]].copy()

    if existing_ph.empty or "scrape_date" not in existing_ph.columns or "unit_id" not in existing_ph.columns:
        return snapshot

    # Latest rent per unit in price_history
    latest = (
        existing_ph
        .sort_values("scrape_date")
        .groupby("unit_id")["rent"]
        .last()
    )
    snapshot["_prev_rent"] = snapshot["unit_id"].map(latest)
    changed = snapshot[
        snapshot["_prev_rent"].isna() |   # new unit
        (snapshot["rent"] != snapshot["_prev_rent"])  # rent changed
    ].drop(columns=["_prev_rent"])
    return changed


# ── Main entry point ───────────────────────────────────────────────────────────

def write_excel(
    scraped: pd.DataFrame,
    registry_df: pd.DataFrame | None,
    output_path: str,
) -> str:
    """
    Append this run's data to the cumulative Excel workbook.

    Args:
        scraped:      Full scraped DataFrame for this run.
        registry_df:  Current unit_registry DataFrame (for the registry sheet).
                      Pass None to skip updating that sheet.
        output_path:  Path to the cumulative .xlsx file.

    Returns:
        output_path
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Load or create workbook
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        logger.info(f"Loaded existing workbook: {output_path}")
    else:
        wb = Workbook()
        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        logger.info(f"Creating new workbook: {output_path}")

    # ── 1. raw_listings ──────────────────────────────────────────────────
    n = _append_new_rows(wb, "raw_listings", scraped, ["unit_id", "scrape_date"])
    logger.info(f"  raw_listings: +{n:,} new rows")

    # ── 2. unit_registry (overwrite) ─────────────────────────────────────
    if registry_df is not None and not registry_df.empty:
        _df_to_sheet_overwrite(wb, "unit_registry", registry_df)
        logger.info(f"  unit_registry: {len(registry_df):,} rows (overwritten)")

    # ── 3. price_history ─────────────────────────────────────────────────
    ph_sheet = "price_history"
    if ph_sheet in wb.sheetnames:
        ws_ph = wb[ph_sheet]
        headers = [ws_ph.cell(1, c).value for c in range(1, ws_ph.max_column + 1)]
        existing_ph_rows = list(ws_ph.iter_rows(min_row=2, values_only=True))
        existing_ph = pd.DataFrame(existing_ph_rows, columns=headers) if existing_ph_rows else pd.DataFrame()
    else:
        existing_ph = pd.DataFrame()

    ph_rows = _build_price_history_rows(scraped, existing_ph)
    n_ph = _append_new_rows(wb, ph_sheet, ph_rows, ["unit_id", "scrape_date"])
    logger.info(f"  price_history: +{n_ph:,} new rows")

    # ── 4. concession_log ─────────────────────────────────────────────────
    concession_cols = ["unit_id", "scrape_date", "reit", "community", "market",
                       "beds", "sqft", "rent", "has_concession", "concession_hardness",
                       "concession_type", "concession_value", "concession_pct_lease_value",
                       "concession_pct_lease_term", "effective_monthly_rent", "concession_raw"]
    concessions = scraped[scraped["has_concession"] == True][[
        c for c in concession_cols if c in scraped.columns
    ]].copy()
    n_c = _append_new_rows(wb, "concession_log", concessions, ["unit_id", "scrape_date"])
    logger.info(f"  concession_log: +{n_c:,} new rows")

    # ── Save ──────────────────────────────────────────────────────────────
    wb.save(output_path)
    logger.info(f"Workbook saved → {output_path}")
    return output_path
